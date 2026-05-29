"""Scripting console execution service.

Maintains a persistent session namespace so variables survive across cell
executions (JupyterLab-style).  Captures stdout/stderr and returns
structured results plus variable inspection.
"""
from __future__ import annotations

import ast
import builtins
import copy
import inspect
import io
import json
import math
import os
import queue
import re
import sys
import threading
import time as py_time
import traceback
import types
from contextlib import redirect_stdout, redirect_stderr
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Set

import pandas as pd

from app_server import config


# ---------------------------------------------------------------------------
# Execution lock & cancellation
# ---------------------------------------------------------------------------

_TIMEOUT_SEC = 60

_DEFAULT_SESSION_ID = "default"
_MAX_SESSION_ID_LEN = 128


class _ScriptTimeout(Exception):
    pass


@dataclass
class _SessionState:
    session_id: str
    namespace: Dict[str, Any] = field(default_factory=dict)
    builtin_keys: set = field(default_factory=set)
    execution_count: int = 0
    exec_lock: threading.Lock = field(default_factory=threading.Lock)
    cancel_event: threading.Event = field(default_factory=threading.Event)
    custom_working_dirs: List[str] = field(default_factory=list)


_SESSION_STATES: Dict[str, _SessionState] = {}
_SESSION_STATES_LOCK = threading.Lock()


def _normalize_session_id(session_id: Optional[str]) -> str:
    sid = (session_id or "").strip()
    if not sid:
        return _DEFAULT_SESSION_ID
    if len(sid) > _MAX_SESSION_ID_LEN:
        sid = sid[:_MAX_SESSION_ID_LEN]
    return sid


def _run_with_timeout(func, timeout_sec: int, cancel_event: threading.Event):
    """Run *func* with a timeout.  Works on Windows (no SIGALRM)."""
    result: Dict[str, Any] = {}
    exc_info: list = [None]

    def _target():
        try:
            result.update(func())
        except BaseException:
            exc_info[0] = sys.exc_info()

    thread = threading.Thread(target=_target, daemon=True)
    thread.start()
    thread.join(timeout=timeout_sec)

    if thread.is_alive():
        # Signal cancellation so the thread can check and exit
        cancel_event.set()
        raise _ScriptTimeout(f"Script exceeded {timeout_sec}s timeout")

    if exc_info[0]:
        raise exc_info[0][1].with_traceback(exc_info[0][2])

    return result


def interrupt_execution(session_id: Optional[str] = None) -> Dict[str, Any]:
    """Signal the running script to stop."""
    session = _get_or_create_session_state(session_id)
    session.cancel_event.set()
    return {"success": True, "message": "Interrupt signal sent."}


def _make_cancel_trace(cancel_event: threading.Event):
    """Create a tracing hook that aborts execution when cancellation is requested."""
    def _trace(_frame, _event, _arg):
        if cancel_event.is_set():
            raise KeyboardInterrupt("Execution cancelled by user")
        return _trace
    return _trace


def _make_interruptible_sleep(
    cancel_event: threading.Event,
    owner_thread_id: int,
    base_sleep_fn,
):
    """Create a thread-scoped sleep that can be interrupted by cancel_event."""
    def _interruptible_sleep(seconds: Any = 0.0):
        # Do not affect other threads that might call time.sleep.
        if threading.get_ident() != owner_thread_id:
            return base_sleep_fn(seconds)

        duration = float(seconds)
        if duration < 0:
            raise ValueError("sleep length must be non-negative")

        deadline = py_time.monotonic() + duration
        while True:
            if cancel_event.is_set():
                raise KeyboardInterrupt("Execution cancelled by user")
            remaining = deadline - py_time.monotonic()
            if remaining <= 0:
                return None
            base_sleep_fn(min(remaining, 0.05))

    return _interruptible_sleep


class _TimeProxy:
    """Proxy that forwards to stdlib time module, except sleep is interruptible."""

    def __init__(self, sleep_fn):
        self.sleep = sleep_fn

    def __getattr__(self, attr: str) -> Any:
        return getattr(py_time, attr)


def _make_session_import_hook(
    base_import,
    cancel_event: threading.Event,
    owner_thread_id: int,
):
    """Wrap __import__ so importing time returns an interruptible proxy."""
    time_proxy = _TimeProxy(
        _make_interruptible_sleep(cancel_event, owner_thread_id, py_time.sleep)
    )

    def _session_import(name, globals=None, locals=None, fromlist=(), level=0):
        module = base_import(name, globals, locals, fromlist, level)
        if str(name).split(".", 1)[0] == "time":
            return time_proxy
        return module

    return _session_import, time_proxy


def _serialize_stream_event(payload: Dict[str, Any]) -> str:
    """Serialize one streaming payload as newline-delimited JSON."""
    return json.dumps(payload, ensure_ascii=False) + "\n"


class _StreamTextSink:
    """Text sink used by redirect_stdout/redirect_stderr with live event emission."""

    def __init__(self, event_queue: "queue.Queue[Dict[str, Any]]", event_type: str):
        self._event_queue = event_queue
        self._event_type = event_type
        self._chunks: List[str] = []
        self.encoding = "utf-8"

    def write(self, text: Any) -> int:
        chunk = str(text or "")
        if not chunk:
            return 0
        self._chunks.append(chunk)
        self._event_queue.put({"type": self._event_type, "text": chunk})
        return len(chunk)

    def flush(self) -> None:
        return None

    def isatty(self) -> bool:
        return False

    def getvalue(self) -> str:
        return "".join(self._chunks)


# ---------------------------------------------------------------------------
# Write-path whitelist
# ---------------------------------------------------------------------------

def _is_write_allowed(path: str, custom_working_dirs: List[str]) -> bool:
    """Check if a file path is in the write whitelist."""
    resolved = os.path.abspath(path)
    allowed_roots = []
    if config.DATA_DIR:
        allowed_roots.append(os.path.abspath(config.DATA_DIR))
    if config.PROJECT_SETTINGS_DIR:
        allowed_roots.append(os.path.abspath(config.PROJECT_SETTINGS_DIR))
    allowed_roots.extend(os.path.abspath(d) for d in custom_working_dirs)
    return any(resolved.startswith(root + os.sep) or resolved == root
               for root in allowed_roots)


def _make_set_working_dir(session: _SessionState):
    def set_working_dir(path: str) -> None:
        """Add a directory to the write whitelist."""
        abs_path = os.path.abspath(path)
        if not os.path.isdir(abs_path):
            raise FileNotFoundError(f"Directory not found: {abs_path}")
        if abs_path not in session.custom_working_dirs:
            session.custom_working_dirs.append(abs_path)
        print(f"Working directory added: {abs_path}")
    return set_working_dir


# ---------------------------------------------------------------------------
# Sandboxed helpers exposed to user scripts
# ---------------------------------------------------------------------------

def _make_read_json():
    def read_json(path: str) -> Any:
        """Read a JSON file and return its contents."""
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return read_json


def _make_write_json(session: _SessionState):
    def write_json(path: str, data: Any, indent: int = 2) -> None:
        """Atomically write data to a JSON file."""
        abs_path = os.path.abspath(path)
        if not _is_write_allowed(abs_path, session.custom_working_dirs):
            raise PermissionError(
                f"Write restricted. Path not in allowed directories: {abs_path}\n"
                f"Use set_working_dir(path) to add a directory to the whitelist."
            )
        os.makedirs(os.path.dirname(abs_path), exist_ok=True)
        tmp = abs_path + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=indent, ensure_ascii=False)
        os.replace(tmp, abs_path)
    return write_json


def _make_read_csv():
    def read_csv(path: str, **kwargs) -> pd.DataFrame:
        """Read a CSV file into a pandas DataFrame."""
        return pd.read_csv(path, **kwargs)
    return read_csv


def _make_write_csv(session: _SessionState):
    def write_csv(path: str, df: pd.DataFrame, index: bool = False, **kwargs) -> None:
        """Atomically write a DataFrame to a CSV file."""
        abs_path = os.path.abspath(path)
        if not _is_write_allowed(abs_path, session.custom_working_dirs):
            raise PermissionError(
                f"Write restricted. Path not in allowed directories: {abs_path}\n"
                f"Use set_working_dir(path) to add a directory to the whitelist."
            )
        os.makedirs(os.path.dirname(abs_path), exist_ok=True)
        tmp = abs_path + ".tmp"
        df.to_csv(tmp, index=index, **kwargs)
        os.replace(tmp, abs_path)
    return write_csv


def _make_list_files():
    def list_files(directory: str, pattern: str = "*") -> List[str]:
        """List files in a directory, optionally filtered by glob pattern."""
        import glob
        return sorted(glob.glob(os.path.join(directory, pattern)))
    return list_files


def _make_get_project_path():
    def get_project_path(project_name: str = "") -> str:
        """Return the base project settings directory, or a specific project folder."""
        base = config.PROJECT_SETTINGS_DIR
        if project_name:
            return os.path.join(base, project_name)
        return base
    return get_project_path


def _make_get_data_path():
    def get_data_path() -> str:
        """Return the data directory path."""
        return config.DATA_DIR
    return get_data_path


def _make_check_cancel(session: _SessionState):
    def check_cancel() -> None:
        """Call in loops to allow cancellation. Raises KeyboardInterrupt if cancelled."""
        if session.cancel_event.is_set():
            raise KeyboardInterrupt("Execution cancelled by user")
    return check_cancel


# ---------------------------------------------------------------------------
# Session namespace (persistent across cell executions)
# ---------------------------------------------------------------------------

def _build_default_namespace(session: _SessionState) -> Dict[str, Any]:
    """Build the default namespace with helper functions and modules."""
    ns: Dict[str, Any] = {
        "__builtins__": __builtins__,
        "read_json": _make_read_json(),
        "write_json": _make_write_json(session),
        "read_csv": _make_read_csv(),
        "write_csv": _make_write_csv(session),
        "list_files": _make_list_files(),
        "get_project_path": _make_get_project_path(),
        "get_data_path": _make_get_data_path(),
        "set_working_dir": _make_set_working_dir(session),
        "check_cancel": _make_check_cancel(session),
        "log": print,
        "pd": pd,
        "json": json,
        "os": os,
        "math": math,
    }
    return ns


def _reset_session_state(session: _SessionState) -> None:
    """Initialize or reset one session namespace."""
    session.custom_working_dirs.clear()
    session.namespace = _build_default_namespace(session)
    session.builtin_keys = set(session.namespace.keys())
    session.execution_count = 0
    session.cancel_event.clear()


def _create_session_state(session_id: str) -> _SessionState:
    session = _SessionState(session_id=session_id)
    _reset_session_state(session)
    return session


def _get_or_create_session_state(session_id: Optional[str]) -> _SessionState:
    sid = _normalize_session_id(session_id)
    with _SESSION_STATES_LOCK:
        session = _SESSION_STATES.get(sid)
        if session is None:
            session = _create_session_state(sid)
            _SESSION_STATES[sid] = session
        return session


# Initialize default session on module load for backward compatibility.
_get_or_create_session_state(_DEFAULT_SESSION_ID)


# ---------------------------------------------------------------------------
# Script execution
# ---------------------------------------------------------------------------


def run_script(code: str, session_id: Optional[str] = None) -> Dict[str, Any]:
    """Execute user Python code in the persistent session namespace."""
    session = _get_or_create_session_state(session_id)

    # Prevent concurrent execution
    if not session.exec_lock.acquire(blocking=False):
        return {
            "success": False,
            "output": "",
            "error": "Another cell is already running. Please wait or interrupt it.",
            "execution_count": session.execution_count,
        }

    try:
        session.execution_count += 1
        session.cancel_event.clear()

        stdout_buf = io.StringIO()
        stderr_buf = io.StringIO()

        def _exec():
            owner_thread_id = threading.get_ident()
            trace_hook = _make_cancel_trace(session.cancel_event)
            previous_builtins = session.namespace.get("__builtins__", __builtins__)
            had_time_binding = "time" in session.namespace
            previous_time_binding = session.namespace.get("time")

            raw_builtins = previous_builtins
            if isinstance(raw_builtins, dict):
                builtins_dict = dict(raw_builtins)
            elif isinstance(raw_builtins, types.ModuleType):
                builtins_dict = dict(vars(raw_builtins))
            else:
                builtins_dict = dict(vars(builtins))

            base_import = builtins_dict.get("__import__", builtins.__import__)
            if not callable(base_import):
                base_import = builtins.__import__

            session_import, time_proxy = _make_session_import_hook(
                base_import,
                session.cancel_event,
                owner_thread_id,
            )
            builtins_dict["__import__"] = session_import

            session.namespace["__builtins__"] = builtins_dict
            session.namespace["time"] = time_proxy

            output = ""
            try:
                with redirect_stdout(stdout_buf), redirect_stderr(stderr_buf):
                    sys.settrace(trace_hook)
                    try:
                        # Parse the code into an AST to check if the last statement
                        # is an expression (like Jupyter: auto-display last expr value)
                        tree = ast.parse(code, "<cell>", "exec")
                        last_expr_value = None

                        if tree.body and isinstance(tree.body[-1], ast.Expr):
                            # Split: exec all statements except last, then eval the last
                            last_node = tree.body.pop()
                            if tree.body:
                                exec(compile(tree, "<cell>", "exec"), session.namespace)
                            # Eval the last expression
                            expr_code = compile(
                                ast.Expression(body=last_node.value), "<cell>", "eval"
                            )
                            last_expr_value = eval(expr_code, session.namespace)
                            # Store as _ (like IPython)
                            session.namespace["_"] = last_expr_value
                        else:
                            exec(compile(tree, "<cell>", "exec"), session.namespace)
                    finally:
                        sys.settrace(None)

                    output = stdout_buf.getvalue()
                    # Append the last expression's repr if it's not None
                    if last_expr_value is not None:
                        expr_repr = repr(last_expr_value)
                        if output and not output.endswith("\n"):
                            output += "\n"
                        output += expr_repr
            finally:
                session.namespace["__builtins__"] = previous_builtins
                if had_time_binding:
                    session.namespace["time"] = previous_time_binding
                else:
                    session.namespace.pop("time", None)

            return {
                "success": True,
                "output": output,
                "error": stderr_buf.getvalue(),
                "execution_count": session.execution_count,
            }

        try:
            result = _run_with_timeout(_exec, _TIMEOUT_SEC, session.cancel_event)
            return result
        except _ScriptTimeout as e:
            return {
                "success": False,
                "output": stdout_buf.getvalue(),
                "error": str(e),
                "execution_count": session.execution_count,
            }
        except KeyboardInterrupt:
            return {
                "success": False,
                "output": stdout_buf.getvalue(),
                "error": "Execution cancelled by user.",
                "execution_count": session.execution_count,
            }
        except SyntaxError as e:
            return {
                "success": False,
                "output": stdout_buf.getvalue(),
                "error": f"SyntaxError: {e.msg} (line {e.lineno})",
                "execution_count": session.execution_count,
            }
        except Exception:
            tb = traceback.format_exc()
            return {
                "success": False,
                "output": stdout_buf.getvalue(),
                "error": tb,
                "execution_count": session.execution_count,
            }
    finally:
        session.cancel_event.clear()
        session.exec_lock.release()


def run_script_stream(code: str, session_id: Optional[str] = None):
    """Execute user code and stream stdout/stderr events as NDJSON lines."""
    session = _get_or_create_session_state(session_id)

    if not session.exec_lock.acquire(blocking=False):
        yield _serialize_stream_event({
            "type": "done",
            "success": False,
            "output": "",
            "error": "Another cell is already running. Please wait or interrupt it.",
            "execution_count": session.execution_count,
        })
        return

    session.execution_count += 1
    session.cancel_event.clear()

    event_queue: "queue.Queue[Dict[str, Any]]" = queue.Queue()
    finished = threading.Event()
    timeout_triggered = threading.Event()
    lock_released = threading.Event()
    started_at = py_time.monotonic()

    stdout_sink = _StreamTextSink(event_queue, "stdout")
    stderr_sink = _StreamTextSink(event_queue, "stderr")
    result_holder: Dict[str, Any] = {
        "success": False,
        "output": "",
        "error": "",
        "execution_count": session.execution_count,
    }

    def _release_exec_lock() -> None:
        if lock_released.is_set():
            return
        lock_released.set()
        session.cancel_event.clear()
        try:
            session.exec_lock.release()
        except RuntimeError:
            # Guard against rare double-release races
            pass

    def _worker():
        nonlocal result_holder
        owner_thread_id = threading.get_ident()
        trace_hook = _make_cancel_trace(session.cancel_event)
        previous_builtins = session.namespace.get("__builtins__", __builtins__)
        had_time_binding = "time" in session.namespace
        previous_time_binding = session.namespace.get("time")

        raw_builtins = previous_builtins
        if isinstance(raw_builtins, dict):
            builtins_dict = dict(raw_builtins)
        elif isinstance(raw_builtins, types.ModuleType):
            builtins_dict = dict(vars(raw_builtins))
        else:
            builtins_dict = dict(vars(builtins))

        base_import = builtins_dict.get("__import__", builtins.__import__)
        if not callable(base_import):
            base_import = builtins.__import__

        session_import, time_proxy = _make_session_import_hook(
            base_import,
            session.cancel_event,
            owner_thread_id,
        )
        builtins_dict["__import__"] = session_import

        session.namespace["__builtins__"] = builtins_dict
        session.namespace["time"] = time_proxy

        try:
            with redirect_stdout(stdout_sink), redirect_stderr(stderr_sink):
                sys.settrace(trace_hook)
                try:
                    tree = ast.parse(code, "<cell>", "exec")
                    last_expr_value = None

                    if tree.body and isinstance(tree.body[-1], ast.Expr):
                        last_node = tree.body.pop()
                        if tree.body:
                            exec(compile(tree, "<cell>", "exec"), session.namespace)
                        expr_code = compile(
                            ast.Expression(body=last_node.value), "<cell>", "eval"
                        )
                        last_expr_value = eval(expr_code, session.namespace)
                        session.namespace["_"] = last_expr_value
                    else:
                        exec(compile(tree, "<cell>", "exec"), session.namespace)
                finally:
                    sys.settrace(None)

                if last_expr_value is not None:
                    expr_repr = repr(last_expr_value)
                    if expr_repr:
                        current_out = stdout_sink.getvalue()
                        if current_out and not current_out.endswith("\n"):
                            stdout_sink.write("\n")
                        stdout_sink.write(expr_repr)

            result_holder = {
                "success": True,
                "output": stdout_sink.getvalue(),
                "error": stderr_sink.getvalue(),
                "execution_count": session.execution_count,
            }
        except KeyboardInterrupt:
            message = (
                f"Script exceeded {_TIMEOUT_SEC}s timeout"
                if timeout_triggered.is_set()
                else "Execution cancelled by user."
            )
            result_holder = {
                "success": False,
                "output": stdout_sink.getvalue(),
                "error": message,
                "execution_count": session.execution_count,
            }
        except SyntaxError as e:
            result_holder = {
                "success": False,
                "output": stdout_sink.getvalue(),
                "error": f"SyntaxError: {e.msg} (line {e.lineno})",
                "execution_count": session.execution_count,
            }
        except Exception:
            result_holder = {
                "success": False,
                "output": stdout_sink.getvalue(),
                "error": traceback.format_exc(),
                "execution_count": session.execution_count,
            }
        finally:
            session.namespace["__builtins__"] = previous_builtins
            if had_time_binding:
                session.namespace["time"] = previous_time_binding
            else:
                session.namespace.pop("time", None)
            finished.set()
            _release_exec_lock()

    worker = threading.Thread(target=_worker, daemon=True)
    worker.start()

    yield _serialize_stream_event({
        "type": "start",
        "execution_count": session.execution_count,
    })

    try:
        while True:
            if (
                not finished.is_set()
                and not timeout_triggered.is_set()
                and (py_time.monotonic() - started_at) > _TIMEOUT_SEC
            ):
                timeout_triggered.set()
                session.cancel_event.set()

            try:
                event = event_queue.get(timeout=0.05)
                yield _serialize_stream_event(event)
                continue
            except queue.Empty:
                pass

            if finished.is_set():
                break

        while True:
            try:
                event = event_queue.get_nowait()
                yield _serialize_stream_event(event)
            except queue.Empty:
                break

        yield _serialize_stream_event({
            "type": "done",
            **result_holder,
        })
    finally:
        # If the client disconnects mid-stream, request cancellation.
        if not finished.is_set():
            session.cancel_event.set()
            worker.join(timeout=0.2)
        _release_exec_lock()


# ---------------------------------------------------------------------------
# Variable inspector
# ---------------------------------------------------------------------------

def _safe_repr(value: Any, max_len: int = 120) -> str:
    """Return a truncated repr of a value."""
    try:
        r = repr(value)
    except Exception:
        r = f"<{type(value).__name__}>"
    if len(r) > max_len:
        r = r[:max_len - 3] + "..."
    return r


def _get_type_label(value: Any) -> str:
    """Return a short, friendly type label."""
    if isinstance(value, pd.DataFrame):
        return "DataFrame"
    if isinstance(value, pd.Series):
        return "Series"
    return type(value).__name__


def _get_size_bytes(value: Any) -> int:
    """Estimate memory usage in bytes."""
    try:
        if isinstance(value, pd.DataFrame):
            return int(value.memory_usage(deep=True).sum())
        if isinstance(value, pd.Series):
            return int(value.memory_usage(deep=True))
        return sys.getsizeof(value)
    except Exception:
        return 0


def _format_size(nbytes: int) -> str:
    """Format byte count as human-readable string."""
    if nbytes < 1024:
        return f"{nbytes} B"
    if nbytes < 1024 * 1024:
        return f"{nbytes / 1024:.1f} KB"
    if nbytes < 1024 * 1024 * 1024:
        return f"{nbytes / (1024 * 1024):.1f} MB"
    return f"{nbytes / (1024 * 1024 * 1024):.1f} GB"


def _get_preview(value: Any) -> str:
    """Return a preview string for the variable inspector."""
    if isinstance(value, pd.DataFrame):
        shape = f"{value.shape[0]} rows x {value.shape[1]} cols"
        cols = ", ".join(str(c) for c in value.columns[:8])
        if len(value.columns) > 8:
            cols += ", ..."
        return f"{shape} [{cols}]"
    if isinstance(value, pd.Series):
        return f"len={len(value)}, dtype={value.dtype}"
    return _safe_repr(value, 100)


def get_variables(session_id: Optional[str] = None) -> List[Dict[str, Any]]:
    """Return user-defined variables from the session namespace."""
    session = _get_or_create_session_state(session_id)
    result: List[Dict[str, Any]] = []
    for key, value in session.namespace.items():
        # Skip built-in namespace items
        if key in session.builtin_keys:
            continue
        # Skip private/dunder names
        if key.startswith("_"):
            continue
        # Skip modules and functions (unless user-defined lambdas)
        if isinstance(value, types.ModuleType):
            continue

        size = _get_size_bytes(value)
        result.append({
            "name": key,
            "type": _get_type_label(value),
            "preview": _get_preview(value),
            "size_bytes": size,
            "size": _format_size(size),
        })

    result.sort(key=lambda v: v["name"])
    return result


def del_variable(name: str, session_id: Optional[str] = None) -> Dict[str, Any]:
    """Delete a user variable from the session namespace."""
    session = _get_or_create_session_state(session_id)
    if name in session.builtin_keys:
        return {"success": False, "message": f"Cannot delete built-in '{name}'."}
    if name not in session.namespace:
        return {"success": False, "message": f"Variable '{name}' not found."}
    del session.namespace[name]
    return {"success": True, "message": f"Variable '{name}' deleted."}


# ---------------------------------------------------------------------------
# Session management
# ---------------------------------------------------------------------------

def reset_session(session_id: Optional[str] = None) -> Dict[str, Any]:
    """Reset the session namespace and execution counter."""
    session = _get_or_create_session_state(session_id)
    if not session.exec_lock.acquire(blocking=False):
        return {"success": False, "message": "Cannot reset while execution is running. Interrupt first."}
    try:
        _reset_session_state(session)
        return {"success": True, "message": "Session reset."}
    finally:
        session.exec_lock.release()


# ---------------------------------------------------------------------------
# Notebook persistence
# ---------------------------------------------------------------------------

def _get_notebooks_dir() -> str:
    """Return the notebooks directory, creating it if needed."""
    preferred = str(getattr(config, "SCRIPTING_DIR", "") or "").strip()
    if preferred:
        nb_dir = preferred
    else:
        nb_dir = os.path.join(config.DATA_DIR, "notebooks") if config.DATA_DIR else os.path.join(os.getcwd(), "notebooks")
    os.makedirs(nb_dir, exist_ok=True)
    return nb_dir


def _sanitize_notebook_filename(filename: str) -> str:
    """Normalize and sanitize user-provided notebook file names."""
    raw = str(filename or "").strip().replace("\\", "/")
    safe = os.path.basename(raw)
    if not safe:
        raise ValueError("Notebook filename is required.")
    return safe


def _normalize_save_filename(filename: str) -> str:
    """Ensure saved notebooks use .ipynb extension."""
    safe = _sanitize_notebook_filename(filename)
    stem, ext = os.path.splitext(safe)
    if ext.lower() == ".ipynb":
        return safe
    base = stem if stem else safe
    return f"{base}.ipynb"


def _source_to_text(source: Any) -> str:
    """Normalize notebook source value to a single string."""
    if isinstance(source, list):
        return "".join(str(part) for part in source)
    if source is None:
        return ""
    return str(source)


def _source_to_lines(source: str) -> List[str]:
    """Convert source text to ipynb-compatible list of lines with newlines."""
    return str(source or "").splitlines(keepends=True)


def _normalize_cell_type(raw_type: Any) -> str:
    """Map notebook cell type to frontend supported set."""
    value = str(raw_type or "").strip().lower()
    if value in {"markdown", "raw"}:
        return value
    return "code"


def _extract_plain_text(data_bundle: Any) -> str:
    """Extract plain text from display data payload."""
    if not isinstance(data_bundle, dict):
        return ""
    return _source_to_text(data_bundle.get("text/plain", ""))


def _convert_outputs_for_import(outputs: Any) -> Dict[str, Any]:
    """Convert ipynb outputs to frontend display fields."""
    if not isinstance(outputs, list):
        return {}

    stdout_parts: List[str] = []
    stderr_parts: List[str] = []
    error_parts: List[str] = []
    unsupported: set[str] = set()

    for item in outputs:
        if not isinstance(item, dict):
            unsupported.add("unknown-output")
            continue

        output_type = str(item.get("output_type", "")).strip().lower()
        if output_type == "stream":
            target = stderr_parts if str(item.get("name", "")).lower() == "stderr" else stdout_parts
            text = _source_to_text(item.get("text", ""))
            if text:
                target.append(text)
            continue

        if output_type == "error":
            traceback_lines = item.get("traceback")
            if isinstance(traceback_lines, list) and traceback_lines:
                error_parts.append("\n".join(str(line) for line in traceback_lines))
            else:
                ename = str(item.get("ename", "Error")).strip()
                evalue = str(item.get("evalue", "")).strip()
                msg = f"{ename}: {evalue}" if evalue else ename
                error_parts.append(msg)
            continue

        if output_type in {"execute_result", "display_data"}:
            data_bundle = item.get("data", {})
            text_plain = _extract_plain_text(data_bundle)
            if text_plain:
                stdout_parts.append(text_plain)

            if isinstance(data_bundle, dict):
                for mime_key in data_bundle.keys():
                    if mime_key != "text/plain":
                        unsupported.add(str(mime_key))
            else:
                unsupported.add(output_type)
            continue

        unsupported.add(output_type or "unknown-output")

    result: Dict[str, Any] = {}
    if stdout_parts:
        result["stdout"] = "".join(stdout_parts)
    if stderr_parts:
        result["stderr"] = "".join(stderr_parts)
    if error_parts:
        result["error"] = "\n".join(error_parts)
    if unsupported:
        result["unsupported"] = sorted(unsupported)
        result["unsupported_message"] = (
            "Imported output contains unsupported rich display types: "
            + ", ".join(sorted(unsupported))
            + "."
        )
    return result


def _normalize_output_text(value: Any) -> str:
    if isinstance(value, list):
        return "".join(str(part) for part in value)
    if value is None:
        return ""
    return str(value)


def _json_safe_value(value: Any) -> Any:
    try:
        json.dumps(value)
        return value
    except (TypeError, ValueError):
        return str(value)


def _normalize_output_data(data: Any) -> Dict[str, Any]:
    if not isinstance(data, dict):
        return {}
    result: Dict[str, Any] = {}
    for key, value in data.items():
        mime_key = str(key or "").strip()
        if not mime_key:
            continue
        result[mime_key] = _json_safe_value(value)
    return result


def _normalize_ipynb_output(output: Any) -> Optional[Dict[str, Any]]:
    if not isinstance(output, dict):
        return None
    output_type = str(output.get("output_type", "")).strip()
    if output_type == "stream":
        name = "stderr" if str(output.get("name", "")).lower() == "stderr" else "stdout"
        text = _normalize_output_text(output.get("text", ""))
        if not text:
            return None
        return {"output_type": "stream", "name": name, "text": text}
    if output_type == "error":
        traceback = output.get("traceback")
        if not isinstance(traceback, list):
            traceback = []
        return {
            "output_type": "error",
            "ename": str(output.get("ename", "Error")),
            "evalue": str(output.get("evalue", "")),
            "traceback": [str(line) for line in traceback],
        }
    if output_type in {"execute_result", "display_data"}:
        result: Dict[str, Any] = {
            "output_type": output_type,
            "data": _normalize_output_data(output.get("data")),
            "metadata": output.get("metadata") if isinstance(output.get("metadata"), dict) else {},
        }
        if output_type == "execute_result":
            execution_count = output.get("execution_count")
            result["execution_count"] = execution_count if isinstance(execution_count, int) else None
        return result
    return None


def _normalize_ipynb_outputs(outputs: Any) -> List[Dict[str, Any]]:
    if not isinstance(outputs, list):
        return []
    result: List[Dict[str, Any]] = []
    for output in outputs:
        normalized = _normalize_ipynb_output(output)
        if normalized:
            result.append(normalized)
    return result


def _to_ipynb_cell(cell: Dict[str, Any]) -> Dict[str, Any]:
    """Convert frontend cell payload to a v4 ipynb cell."""
    cell_type = _normalize_cell_type(cell.get("type"))
    source_text = _source_to_text(cell.get("source", ""))
    base: Dict[str, Any] = {
        "cell_type": cell_type,
        "metadata": {},
        "source": _source_to_lines(source_text),
    }
    if cell_type == "code":
        execution_count = cell.get("execution_count")
        base["execution_count"] = execution_count if isinstance(execution_count, int) else None
        base["outputs"] = _normalize_ipynb_outputs(cell.get("outputs"))
    return base


def _from_ipynb_cell(cell: Dict[str, Any]) -> Dict[str, Any]:
    """Convert one ipynb cell to frontend cell payload."""
    cell_type = _normalize_cell_type(cell.get("cell_type"))
    source = _source_to_text(cell.get("source", ""))
    frontend_cell: Dict[str, Any] = {
        "type": cell_type,
        "source": source,
    }
    if cell_type == "code":
        execution_count = cell.get("execution_count")
        if isinstance(execution_count, int):
            frontend_cell["execution_count"] = execution_count
        outputs = _normalize_ipynb_outputs(cell.get("outputs"))
        frontend_cell["outputs"] = outputs
        output_info = _convert_outputs_for_import(outputs)
        if output_info:
            frontend_cell["import_output"] = output_info
    return frontend_cell


def _load_arcnb_cells(data: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Load legacy .arcnb files and normalize to frontend cells."""
    raw_cells = data.get("cells", [])
    if not isinstance(raw_cells, list):
        return []
    result: List[Dict[str, Any]] = []
    for entry in raw_cells:
        if not isinstance(entry, dict):
            continue
        result.append({
            "type": _normalize_cell_type(entry.get("type")),
            "source": _source_to_text(entry.get("source", "")),
        })
    return result


def save_notebook(filename: str, cells: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Save cells to a .ipynb notebook file."""
    try:
        filename = _normalize_save_filename(filename)
    except ValueError as exc:
        return {"success": False, "message": str(exc)}
    nb_dir = _get_notebooks_dir()
    filepath = os.path.join(nb_dir, filename)
    data: Dict[str, Any] = {
        "cells": [_to_ipynb_cell(c if isinstance(c, dict) else {}) for c in (cells or [])],
        "metadata": {
            "kernelspec": {
                "display_name": "Python 3",
                "language": "python",
                "name": "python3",
            },
            "language_info": {
                "name": "python",
            },
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    tmp = filepath + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp, filepath)
    return {"success": True, "path": filepath, "message": f"Saved to {filename}"}


def load_notebook(filename: str) -> Dict[str, Any]:
    """Load cells from a notebook file (.ipynb preferred, .arcnb legacy)."""
    nb_dir = _get_notebooks_dir()
    try:
        safe_name = _sanitize_notebook_filename(filename)
    except ValueError as exc:
        return {"success": False, "message": str(exc), "cells": []}
    name_stem, name_ext = os.path.splitext(safe_name)
    candidates: List[str] = []
    if name_ext:
        candidates.append(safe_name)
    else:
        candidates.extend([f"{safe_name}.ipynb", f"{safe_name}.arcnb"])
    if name_ext.lower() == ".ipynb":
        candidates.append(f"{name_stem}.arcnb")

    filepath = ""
    for candidate in candidates:
        candidate_path = os.path.join(nb_dir, candidate)
        if os.path.isfile(candidate_path):
            filepath = candidate_path
            break

    if not filepath:
        requested = safe_name if safe_name else filename
        return {"success": False, "message": f"File not found: {requested}", "cells": []}

    if not os.path.isfile(filepath):
        return {"success": False, "message": f"File not found: {safe_name}", "cells": []}
    with open(filepath, "r", encoding="utf-8") as f:
        data = json.load(f)
    _, ext = os.path.splitext(filepath)
    if ext.lower() == ".ipynb":
        raw_cells = data.get("cells", [])
        if not isinstance(raw_cells, list):
            raw_cells = []
        cells = [_from_ipynb_cell(c) for c in raw_cells if isinstance(c, dict)]
    else:
        cells = _load_arcnb_cells(data)
    return {"success": True, "cells": cells, "path": filepath}


def list_notebooks() -> List[Dict[str, str]]:
    """List available notebook files."""
    nb_dir = _get_notebooks_dir()
    result = []
    for entry in sorted(os.listdir(nb_dir)):
        lower = entry.lower()
        if lower.endswith(".ipynb") or lower.endswith(".arcnb"):
            filepath = os.path.join(nb_dir, entry)
            stat = os.stat(filepath)
            result.append({
                "name": entry,
                "size": _format_size(stat.st_size),
                "modified": str(int(stat.st_mtime)),
            })
    return result


# ---------------------------------------------------------------------------
# User macros
# ---------------------------------------------------------------------------

_MACRO_META_BEGIN = "# <arcrho-macro>"
_MACRO_META_END = "# </arcrho-macro>"
_APPLY_ADJUSTMENTS_MACRO = r'''# <arcrho-macro>
# title: Apply Growth Adjustments
# description: Reads COL growth adjustments from the workbook used by the 2026Q1 reserve-review notebook, applies them to the active DFM method, and adds adjustment notes.
# </arcrho-macro>

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover - runtime user dependency
    load_workbook = None
    _LOAD_WORKBOOK_ERROR = exc
else:
    _LOAD_WORKBOOK_ERROR = None

try:
    from arcrho_api.exceptions import DfmDataError
except Exception:  # pragma: no cover - script can still show useful errors
    DfmDataError = ValueError


# Match the legacy notebook style:
# sheet = load_workbook(file_path, data_only=1)["Summary"]
# adjustment = {"incurred loss": [sheet[f"F{i}"].value for i in [7, 8, 9]], ...}
GROWTH_ADJUSTMENT_WORKBOOK = (
    r"E:\ResQ\Automations\Reserve Review\2026Q1\Growth Adjustment 2026Q1.xlsx"
)
GROWTH_ADJUSTMENT_SHEET = "Summary"
PRE_ADJUSTMENT_CELL_NOTE = "Selected before adjustments."


def load_adjustments_from_workbook(file_path: str = GROWTH_ADJUSTMENT_WORKBOOK) -> dict[str, list[Any]]:
    if load_workbook is None:
        raise DfmDataError(f"openpyxl is required to load adjustment workbooks: {_LOAD_WORKBOOK_ERROR}")
    workbook_path = Path(file_path)
    if not workbook_path.exists():
        raise DfmDataError(f"Growth adjustment workbook was not found: {workbook_path}")
    sheet = load_workbook(str(workbook_path), data_only=1)[GROWTH_ADJUSTMENT_SHEET]
    return {
        "incurred loss": [sheet[f"F{i}"].value for i in [7, 8, 9]],
        "paid loss": [sheet[f"F{i}"].value for i in [10, 11, 12]],
        "counts": [sheet[f"F{i}"].value for i in [13, 14, 15]],
        "accounting cutoff": [sheet["C20"].value],
        "other": [0],
    }


def _clean_text(value: Any) -> str:
    return " ".join(str(value if value is not None else "").split()).strip()


def _normalize_label(value: Any) -> str:
    return _clean_text(value)


def _label_key(value: Any) -> str:
    label = _normalize_label(value)
    if ":" in label:
        prefix, rest = label.split(":", 1)
        if prefix.strip().isdigit():
            label = rest.strip()
    return label.lower()


def _number(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number != number or number in (float("inf"), float("-inf")):
        return None
    return number


def _adjustment_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None
        try:
            if text.endswith("%"):
                return round(float(text[:-1].strip()) / 100, 4)
            return float(text)
        except ValueError as err:
            raise DfmDataError(f"Invalid adjustment value: {value!r}") from err
    number = _number(value)
    if number is None:
        return None
    return round(number, 4)


def _ensure_matrix(container: dict[str, Any], key: str, rows: int, cols: int, fill: Any = 0) -> list[list[Any]]:
    existing = container.get(key)
    if not isinstance(existing, list):
        existing = []
    existing = [row if isinstance(row, list) else [] for row in existing]
    while len(existing) < rows:
        existing.append([])
    for row in existing:
        while len(row) < cols:
            row.append(fill)
    existing = existing[:rows]
    for index, row in enumerate(existing):
        existing[index] = row[:cols]
    container[key] = existing
    return existing


def _normalize_adjustment_list(values: Any) -> list[float]:
    if values is None:
        return []
    if isinstance(values, (str, bytes)):
        source = [values]
    else:
        try:
            source = list(values)
        except TypeError:
            source = [values]
    out: list[float] = []
    for value in source:
        number = _adjustment_number(value)
        if number is not None:
            out.append(number)
    return out


def _normalize_adjustments(adjustments: Any) -> dict[str, list[float]]:
    if not isinstance(adjustments, dict):
        raise DfmDataError("apply_adjustments requires adjustments to be a dictionary.")
    aliases = {
        "count": "counts",
        "claim count": "counts",
        "claim counts": "counts",
        "paid": "paid loss",
        "paid losses": "paid loss",
        "incurred": "incurred loss",
        "incurred losses": "incurred loss",
        "accounting": "accounting cutoff",
        "cutoff": "accounting cutoff",
    }
    out: dict[str, list[float]] = {
        "counts": [],
        "paid loss": [],
        "incurred loss": [],
        "accounting cutoff": [],
        "other": [],
    }
    for key, values in adjustments.items():
        normalized_key = aliases.get(_clean_text(key).lower(), _clean_text(key).lower())
        if normalized_key in out:
            out[normalized_key] = _normalize_adjustment_list(values)
    return out


def _adjustment_method_kind(dfm: Any) -> str:
    try:
        info = dfm.output_vector_dataset_type()
    except Exception:
        info = None
    if info is not None:
        dataset_type = _clean_text(getattr(info, "name", ""))
        category = _clean_text(getattr(info, "category", ""))
        dataset_type_lower = dataset_type.lower()
        category_lower = category.lower()
        if "52" in dataset_type:
            return "skip"
        if category_lower == "c claim count":
            return "counts"
        if category_lower == "h severity":
            return "severity"
        if "paid" in dataset_type_lower or "salv dfm" in dataset_type_lower or "subr dfm" in dataset_type_lower:
            return "paid"
        if "incurred" in dataset_type_lower:
            return "incurred"

    text = f"{dfm.output_vector} {dfm.name}".lower()
    if re.search(r"(^|\D)52(\D|$)", text):
        return "skip"
    if re.search(r"(^|\W)h\s*\d+", text) or "severity" in text:
        return "severity"
    if "paid" in text or "salv" in text or "subr" in text:
        return "paid"
    if "incurred" in text:
        return "incurred"
    if re.search(r"(^|\W)c\s*\d+", text) or "claim count" in text or "counts" in text:
        return "counts"
    if "reported" in text or "cwp" in text or "cwop" in text:
        return "counts"
    return "incurred"


def _selected_average_row(selected: list[list[Any]], col: int) -> int | None:
    for row, row_values in enumerate(selected):
        if col < len(row_values) and bool(row_values[col]):
            return row
    return None


def _format_adjustment_percent(value: float) -> str:
    return f"{round(abs(value) * 100, 2):g}%"


def _factor_note_part(value: float) -> str:
    if value > 0:
        return f"1+{_format_adjustment_percent(value)}"
    if value < 0:
        return f"1-{_format_adjustment_percent(value)}"
    return "1"


def _compound_adjustment_part(values: list[float], col: int, *, formula_style: str) -> dict[str, Any]:
    current = values[col] if col < len(values) else 0.0
    next_value = values[col + 1] if col + 1 < len(values) else 0.0
    current_factor = 1.0 + current
    next_factor = 1.0 + next_value
    if col + 1 < len(values):
        value = current_factor / next_factor if next_factor else current_factor
    else:
        value = current_factor
    if formula_style == "left":
        formula = _factor_note_part(current)
        if col + 1 < len(values) and next_value != 0:
            formula = f"{formula}/({_factor_note_part(next_value)})"
        if formula.count("(") == 1 and "/" not in formula:
            formula = formula.replace("(", "").replace(")", "")
    else:
        if col + 1 < len(values) and next_factor != 1:
            formula = f"{round(current_factor, 4):g}/{round(next_factor, 4):g}"
        else:
            formula = f"{round(current_factor, 4):g}"
    return {"value": value, "formula": formula}


def _adjustment_for_kind(kind: str, adjustments: dict[str, list[float]], col: int) -> dict[str, Any]:
    counts_left = _compound_adjustment_part(adjustments["counts"], col, formula_style="left")
    counts_right = _compound_adjustment_part(adjustments["counts"], col, formula_style="right")
    incurred_left = _compound_adjustment_part(adjustments["incurred loss"], col, formula_style="left")
    incurred_right = _compound_adjustment_part(adjustments["incurred loss"], col, formula_style="right")
    paid_left = _compound_adjustment_part(adjustments["paid loss"], col, formula_style="left")
    paid_right = _compound_adjustment_part(adjustments["paid loss"], col, formula_style="right")
    if kind == "counts":
        return {"factor": counts_right["value"], "left": counts_left["formula"], "right": counts_right["formula"]}
    if kind == "paid":
        return {"factor": paid_right["value"], "left": paid_left["formula"], "right": paid_right["formula"]}
    if kind == "severity":
        factor = incurred_right["value"] / counts_right["value"] if counts_right["value"] else 1.0
        left = f"{incurred_left['formula']}/{counts_left['formula']}"
        right = f"({incurred_right['formula']})/({counts_right['formula']})"
        return {"factor": factor, "left": left, "right": right}
    return {"factor": incurred_right["value"], "left": incurred_left["formula"], "right": incurred_right["formula"]}


def _has_meaningful_adjustment(growth: float, accounting_cutoff: float, other_factor: float) -> bool:
    return any(abs(value - 1.0) > 0.0000001 for value in (growth, accounting_cutoff, other_factor))


def _display_average_label(label: str) -> str:
    text = _normalize_label(label)
    if ":" in text:
        _prefix, rest = text.split(":", 1)
        if rest.strip():
            return rest.strip()
    return text


def _mark_selected_before_adjustment(dfm: Any, col: int, label: str) -> None:
    dfm.clear_cell_notes_for_development(col + 1)
    dfm.set_cell_note(_display_average_label(label), col + 1, PRE_ADJUSTMENT_CELL_NOTE)


def _format_adjustment_note(
    dfm: Any,
    col: int,
    label: str,
    average_value: float,
    final_value: float,
    adjustment: dict[str, Any],
    accounting_cutoff: float,
    other_factor: float,
) -> str:
    lines = [f"For development period {dfm.dev_period(col + 1)}:"]
    formula_parts = [f"{average_value:.4f}"]
    if abs(float(adjustment["factor"]) - 1.0) > 0.0000001:
        lines.append(f"  - Apply growth adjustments of {adjustment['left']} = {adjustment['right']};")
        formula_parts.append(str(adjustment["right"]))
    if abs(accounting_cutoff - 1.0) > 0.0000001:
        lines.append(f"  - Apply accounting cutoff 1+{accounting_cutoff - 1.0:.2%} = {accounting_cutoff:.4f};")
        formula_parts.append(f"{accounting_cutoff:.4f}")
    if abs(other_factor - 1.0) > 0.0000001:
        formula_parts.append(f"{other_factor:.4f}")
    display_label = _display_average_label(label)
    lines.append(f"  - Selected average factor: \"{display_label}\" ({average_value:.4f})")
    lines.append(f"  - Selected LDF after adjustments: {' * '.join(formula_parts)} = {final_value:.4f}")
    return "\n".join(lines)


def _clear_adjustment_notes(dfm: Any) -> None:
    keywords = (
        "For development period ",
        "Apply growth adjustments of ",
        "Apply accounting cutoff ",
        "Selected average factor: ",
        "Selected LDF after adjustments: ",
    )
    lines = [line for line in dfm.notes.splitlines() if not any(keyword in line for keyword in keywords)]
    dfm.update_notes("\n".join(lines).strip())


def apply_adjustments(
    dfm: Any,
    selection: str | None = None,
    *,
    adjustments: dict[str, Any],
    other_adjustment: Any = None,
    add_notes: bool = True,
    clear_prior_notes: bool = True,
) -> Any:
    if selection:
        dfm.set_selected_average(selection)

    normalized = _normalize_adjustments(adjustments)
    other_values = _normalize_adjustment_list(
        other_adjustment if other_adjustment is not None else normalized.get("other", [0])
    )
    dev_count = max(
        [len(values) for values in normalized.values()] + [len(other_values), dfm._average_col_count()],
        default=0,
    )
    if dev_count <= 0:
        return dfm
    for key in ("counts", "paid loss", "incurred loss", "accounting cutoff", "other"):
        values = normalized.setdefault(key, [])
        if len(values) < dev_count:
            values.extend([0.0] * (dev_count - len(values)))
    if len(other_values) < dev_count:
        other_values.extend([0.0] * (dev_count - len(other_values)))

    method_kind = _adjustment_method_kind(dfm)
    if method_kind == "skip":
        return dfm

    labels = dfm._average_labels()
    selected = _ensure_matrix(dfm.average_formulas, "selected", len(labels), dfm._average_col_count(), 0)
    values = _ensure_matrix(dfm.average_formulas, "values", len(labels), dfm._average_col_count(), None)
    if clear_prior_notes:
        _clear_adjustment_notes(dfm)

    changed = False
    note_blocks: list[str] = []
    for col in range(min(dev_count, dfm._average_col_count())):
        selected_row = _selected_average_row(selected, col)
        if selected_row is None:
            if selection:
                selected_row = dfm._ensure_average_label(selection)
                labels = dfm._average_labels()
                selected = _ensure_matrix(dfm.average_formulas, "selected", len(labels), dfm._average_col_count(), 0)
                values = _ensure_matrix(dfm.average_formulas, "values", len(labels), dfm._average_col_count(), None)
            else:
                continue
        if selected_row >= len(labels):
            continue
        average_value = _number(values[selected_row][col] if col < len(values[selected_row]) else None)
        if average_value is None:
            continue
        adjustment = _adjustment_for_kind(method_kind, normalized, col)
        accounting_cutoff = 1.0 if method_kind == "severity" else 1.0 + normalized["accounting cutoff"][col]
        other_factor = 1.0 + other_values[col]
        final_value = average_value * adjustment["factor"] * accounting_cutoff * other_factor
        if not _has_meaningful_adjustment(adjustment["factor"], accounting_cutoff, other_factor):
            continue
        _mark_selected_before_adjustment(dfm, col, labels[selected_row])
        dfm.set_user_ratio(round(final_value, 4), col + 1)
        changed = True
        if add_notes:
            note_blocks.append(_format_adjustment_note(
                dfm,
                col,
                labels[selected_row],
                average_value,
                final_value,
                adjustment,
                accounting_cutoff,
                other_factor,
            ))

    if add_notes and note_blocks:
        existing = dfm.notes
        suffix = "\n\n".join(note_blocks)
        dfm.update_notes(f"{existing}\n\n{suffix}" if existing else suffix)
    if not changed and add_notes:
        dfm.add_notes("No growth/accounting cutoff adjustments were needed for this method.")
    return dfm


def run_macro(active_dfm, active_context=None):
    adjustment = load_adjustments_from_workbook()
    apply_adjustments(active_dfm, adjustments=adjustment)
    return active_dfm
'''


def _get_macros_dir() -> str:
    macro_dir = _get_notebooks_dir()
    os.makedirs(macro_dir, exist_ok=True)
    return macro_dir


def _seed_builtin_macros() -> None:
    macro_dir = _get_macros_dir()
    target = os.path.join(macro_dir, "apply_growth_adjustments.py")
    if not os.path.exists(target):
        tmp = target + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            f.write(_APPLY_ADJUSTMENTS_MACRO)
        os.replace(tmp, target)


def _parse_macro_metadata(text: str, filename: str) -> Dict[str, str]:
    title = os.path.splitext(os.path.basename(filename))[0].replace("_", " ").title()
    description = ""
    in_block = False
    for raw_line in str(text or "").splitlines():
        line = raw_line.strip()
        if line == _MACRO_META_BEGIN:
            in_block = True
            continue
        if line == _MACRO_META_END:
            break
        if not in_block:
            continue
        if line.startswith("#"):
            line = line[1:].strip()
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        key = key.strip().lower()
        value = value.strip()
        if key == "title" and value:
            title = value
        elif key == "description" and value:
            description = value
    return {"title": title, "description": description}


def _safe_macro_path(macro_id: str) -> str:
    macro_dir = os.path.abspath(_get_macros_dir())
    safe_name = os.path.basename(str(macro_id or "").strip().replace("\\", "/"))
    if not safe_name:
        raise ValueError("Macro id is required.")
    if not safe_name.lower().endswith(".py"):
        safe_name = f"{safe_name}.py"
    path = os.path.abspath(os.path.join(macro_dir, safe_name))
    if not (path == macro_dir or path.startswith(macro_dir + os.sep)):
        raise ValueError("Macro path is outside the scripting directory.")
    return path


def list_macros() -> List[Dict[str, str]]:
    _seed_builtin_macros()
    macro_dir = _get_macros_dir()
    result: List[Dict[str, str]] = []
    for entry in sorted(os.listdir(macro_dir)):
        if not entry.lower().endswith(".py"):
            continue
        path = os.path.join(macro_dir, entry)
        if not os.path.isfile(path):
            continue
        try:
            text = Path(path).read_text(encoding="utf-8")
            meta = _parse_macro_metadata(text, entry)
            stat = os.stat(path)
            result.append({
                "id": entry,
                "name": meta["title"],
                "description": meta["description"],
                "path": path,
                "modified": str(int(stat.st_mtime)),
            })
        except OSError:
            continue
    return result


def _ensure_arcrho_api_import_path() -> None:
    repo_root = Path(__file__).resolve().parents[3]
    api_src = repo_root / "python-api" / "src"
    if api_src.exists():
        text = str(api_src)
        if text not in sys.path:
            sys.path.insert(0, text)


def _runtime_active_dfm_path() -> str:
    runtime_dir = os.path.join(_get_macros_dir(), ".macro_runtime")
    os.makedirs(runtime_dir, exist_ok=True)
    return os.path.join(runtime_dir, "active-dfm.json")


def _build_active_dfm(active_context: Dict[str, Any]):
    _ensure_arcrho_api_import_path()
    from arcrho_api import ArcRhoClient, DfmMethod

    active_json = active_context.get("activeJson")
    if not isinstance(active_json, dict):
        raise ValueError("Active DFM JSON is not available.")
    active_json = copy.deepcopy(active_json)
    fields = active_context.get("fields") if isinstance(active_context.get("fields"), dict) else {}
    details = active_json.get("details tab") if isinstance(active_json.get("details tab"), dict) else {}
    metadata = active_json.get("method metadata") if isinstance(active_json.get("method metadata"), dict) else {}

    project_name = str(fields.get("project") or metadata.get("project") or "").strip()
    reserving_class = str(fields.get("reservingClass") or details.get("reserving class") or "").strip()
    method_name = str(fields.get("methodName") or details.get("name") or "").strip()
    method_path = str(active_context.get("methodPath") or "").strip()

    dfm = None
    if project_name and reserving_class and method_name:
        try:
            dfm = ArcRhoClient(config.get_root_path()).project(project_name).reserving_class(reserving_class).dfm(method_name)
        except Exception:
            dfm = None
    if dfm is not None:
        dfm.payload = active_json
        dfm._ensure_grouped_payload()
        if method_path:
            dfm.file_path = Path(method_path)
        return dfm

    temp_path = _runtime_active_dfm_path()
    with open(temp_path, "w", encoding="utf-8") as f:
        json.dump(active_json, f, indent=2, ensure_ascii=False)
    return DfmMethod.load_file(temp_path)


def run_macro(macro_id: str, active_context: Dict[str, Any]) -> Dict[str, Any]:
    _seed_builtin_macros()
    path = _safe_macro_path(macro_id)
    if not os.path.isfile(path):
        return {"success": False, "message": f"Macro not found: {macro_id}"}
    try:
        source = Path(path).read_text(encoding="utf-8")
        active_dfm = _build_active_dfm(active_context if isinstance(active_context, dict) else {})
        output = io.StringIO()
        namespace: Dict[str, Any] = {
            "__name__": "__arcrho_macro__",
            "active_dfm": active_dfm,
            "dfm": active_dfm,
            "active_context": active_context,
            "log": print,
        }
        with redirect_stdout(output):
            exec(compile(source, path, "exec"), namespace)
            runner = namespace.get("run_macro") or namespace.get("main")
            if callable(runner):
                runner(active_dfm, active_context)
        return {
            "success": True,
            "message": f"Ran {os.path.basename(path)}",
            "payload": active_dfm.to_dict(),
            "stdout": output.getvalue(),
            "path": path,
        }
    except Exception as exc:
        return {
            "success": False,
            "message": str(exc),
            "traceback": traceback.format_exc(),
            "path": path,
        }


# ---------------------------------------------------------------------------
# API documentation
# ---------------------------------------------------------------------------

def get_api_help() -> List[Dict[str, str]]:
    """Return documentation for functions available in scripts."""
    return [
        {"name": "read_json(path)", "description": "Read a JSON file and return its contents as a Python object."},
        {"name": "write_json(path, data, indent=2)", "description": "Atomically write data to a JSON file. Path must be in an allowed directory."},
        {"name": "read_csv(path, **kwargs)", "description": "Read a CSV file into a pandas DataFrame."},
        {"name": "write_csv(path, df, index=False, **kwargs)", "description": "Atomically write a DataFrame to a CSV file. Path must be in an allowed directory."},
        {"name": "list_files(directory, pattern='*')", "description": "List files in a directory, optionally filtered by glob pattern."},
        {"name": "get_project_path(project_name='')", "description": "Return the project settings directory, or a specific project folder."},
        {"name": "get_data_path()", "description": "Return the data directory path."},
        {"name": "set_working_dir(path)", "description": "Add a directory to the write whitelist so write_json/write_csv can write to it."},
        {"name": "check_cancel()", "description": "Call inside long loops to allow user cancellation. Raises KeyboardInterrupt if cancelled."},
        {"name": "log(message)", "description": "Print a message to the output console."},
        {"name": "pd", "description": "The pandas module, available as 'pd'."},
        {"name": "json", "description": "The json standard library module."},
        {"name": "os", "description": "The os standard library module."},
        {"name": "math", "description": "The math standard library module."},
    ]


# ---------------------------------------------------------------------------
# User preferences
# ---------------------------------------------------------------------------

_SCRIPTING_PREFS_LOCK = threading.Lock()
_LOCAL_PROJECT_PREFS_LOCK = threading.Lock()
_LEGACY_DATASET_VIEWER_PREFS_KEY = "dataset_viewer_local_prefs_v1"


def get_preferences() -> Dict[str, Any]:
    """Load scripting user preferences from APPDATA JSON file."""
    filepath = config.get_scripting_prefs_path()
    with _SCRIPTING_PREFS_LOCK:
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                return data
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            pass
    return {}


def save_preferences(prefs: Dict[str, Any]) -> Dict[str, Any]:
    """Merge and save scripting user preferences to APPDATA JSON file."""
    filepath = config.get_scripting_prefs_path()
    with _SCRIPTING_PREFS_LOCK:
        # Load existing, merge with incoming
        existing: Dict[str, Any] = {}
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            if isinstance(loaded, dict):
                existing = loaded
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            pass

        existing.update(prefs)

        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(existing, f, indent=2, ensure_ascii=False)

    return {"success": True, "preferences": existing}


def _normalize_local_project_preferences(raw: Any) -> Dict[str, Any]:
    source = raw if isinstance(raw, dict) else {}
    project = str(
        source.get("projectName")
        or source.get("project_name")
        or source.get("project")
        or ""
    ).strip()
    updated_at = str(source.get("updated_at") or source.get("updatedAt") or "").strip()
    out: Dict[str, Any] = {}
    if project:
        out["projectName"] = project
    recent_raw = (
        source.get("recentProjectNames")
        or source.get("recent_project_names")
        or source.get("recentProjects")
        or source.get("recent_projects")
        or []
    )
    if isinstance(recent_raw, (list, tuple)):
        recent_projects: List[str] = []
        seen_projects: Set[str] = set()
        for item in recent_raw:
            recent_project = str(item or "").strip()
            recent_key = recent_project.lower()
            if not recent_project or recent_key in seen_projects:
                continue
            seen_projects.add(recent_key)
            recent_projects.append(recent_project)
            if len(recent_projects) >= 3:
                break
        if recent_projects:
            out["recentProjectNames"] = recent_projects
    if updated_at:
        out["updated_at"] = updated_at

    explorer_source = None
    for explorer_key in (
        "projectExplorer",
        "project_explorer",
        "projectSettingsExplorer",
        "project_settings_explorer",
    ):
        if explorer_key in source:
            explorer_source = source.get(explorer_key)
            break
    if isinstance(explorer_source, dict):
        expanded_raw = (
            explorer_source.get("expandedFolders")
            or explorer_source.get("expanded_folders")
            or []
        )
        if isinstance(expanded_raw, (list, tuple)):
            expanded_folders: List[str] = []
            seen_folders: Set[str] = set()
            for item in expanded_raw:
                folder = str(item or "").strip().replace("/", "\\")
                folder = "\\".join(part.strip() for part in folder.split("\\") if part.strip())
                folder_key = folder.lower()
                if not folder or folder_key in seen_folders:
                    continue
                seen_folders.add(folder_key)
                expanded_folders.append(folder)
            out["projectExplorer"] = {"expandedFolders": expanded_folders}
    return out


def get_local_project_preferences() -> Dict[str, Any]:
    """Load shared last-project preferences from a dedicated APPDATA JSON file."""
    filepath = config.get_local_project_prefs_path()
    with _LOCAL_PROJECT_PREFS_LOCK:
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                data = json.load(f)
            normalized = _normalize_local_project_preferences(data)
            if normalized:
                return normalized
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            pass

    legacy = _normalize_local_project_preferences(
        get_preferences().get(_LEGACY_DATASET_VIEWER_PREFS_KEY)
    )
    return legacy


def save_local_project_preferences(prefs: Dict[str, Any]) -> Dict[str, Any]:
    """Merge shared last-project preferences into %APPDATA%\\ArcRho\\local_project_prefs.json."""
    filepath = config.get_local_project_prefs_path()
    with _LOCAL_PROJECT_PREFS_LOCK:
        existing: Dict[str, Any] = {}
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                loaded = json.load(f)
            existing = _normalize_local_project_preferences(loaded)
        except (FileNotFoundError, json.JSONDecodeError, OSError):
            existing = {}

        incoming = _normalize_local_project_preferences(prefs)
        incoming_source = prefs if isinstance(prefs, dict) else {}
        incoming_has_recent_projects = any(
            key in incoming_source
            for key in ("recentProjectNames", "recent_project_names", "recentProjects", "recent_projects")
        )
        incoming_project = str(incoming.get("projectName") or "").strip()
        existing_recent = existing.get("recentProjectNames")
        incoming_recent = incoming.get("recentProjectNames")
        merged_recent: List[str] = []
        if incoming_has_recent_projects:
            seen_recent: Set[str] = set()
            for candidate in [
                incoming_project,
                *(incoming_recent if isinstance(incoming_recent, list) else []),
                *(existing_recent if isinstance(existing_recent, list) else []),
            ]:
                recent_project = str(candidate or "").strip()
                recent_key = recent_project.lower()
                if not recent_project or recent_key in seen_recent:
                    continue
                seen_recent.add(recent_key)
                merged_recent.append(recent_project)
                if len(merged_recent) >= 3:
                    break

        existing.update(incoming)
        if incoming_has_recent_projects and merged_recent:
            existing["recentProjectNames"] = merged_recent

        os.makedirs(os.path.dirname(filepath), exist_ok=True)
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(existing, f, indent=2, ensure_ascii=False)
            f.write("\n")

    return {"success": True, "preferences": existing}


# ---------------------------------------------------------------------------
# Object introspection (Shift+Tab tooltip)
# ---------------------------------------------------------------------------

_IDENT_RE = re.compile(r"[A-Za-z_]\w*(?:\.[A-Za-z_]\w*)*")
_MAX_DOC_LEN = 2000
_MAX_DICT_KEYS = 20
_MAX_DF_COLS = 30


def _extract_expression_at(code: str, cursor_pos: int) -> List[str]:
    """Extract candidate dotted identifier expressions around *cursor_pos*.

    Returns a list of candidates ordered by priority:
    1. The direct identifier under/adjacent to the cursor
    2. The callable name before the enclosing ``(`` if cursor is inside parens

    The caller should try each candidate in order until one resolves.
    """
    pos = max(0, min(cursor_pos, len(code)))
    candidates: List[str] = []

    # --- Candidate 1: direct identifier at cursor ---
    start = pos
    while start > 0 and (code[start - 1].isalnum() or code[start - 1] in ("_", ".")):
        start -= 1
    end = pos
    while end < len(code) and (code[end].isalnum() or code[end] == "_"):
        end += 1
    fragment = code[start:end].strip(".")
    if fragment and _IDENT_RE.fullmatch(fragment):
        candidates.append(fragment)

    # --- Candidate 2: callable before enclosing '(' ---
    # Walk backwards, skipping over string literals to handle cases like
    # func(r"E:\path\file.csv") where cursor is inside the string.
    depth = 0
    i = pos - 1
    in_str: Optional[str] = None  # track if we're inside a string (scanning backwards)
    while i >= 0:
        ch = code[i]
        # Simple backwards string skipping: when we hit a quote, scan back
        # to find its matching opening quote.
        if in_str is None and ch in ('"', "'"):
            quote = ch
            # Check for triple-quote
            if i >= 2 and code[i - 2:i + 1] == quote * 3:
                j = i - 3
                triple = quote * 3
                idx = code.rfind(triple, 0, j + 1)
                if idx >= 0:
                    i = idx - 1
                else:
                    break
                continue
            # Single quote — find matching opening quote (skip escaped quotes)
            j = i - 1
            while j >= 0:
                if code[j] == quote and (j == 0 or code[j - 1] != "\\"):
                    break
                j -= 1
            # Also handle raw-string prefix: r", b", etc.
            if j > 0 and code[j - 1] in ("r", "R", "b", "B", "f", "F", "u", "U"):
                j -= 1
            if j > 1 and code[j - 1:j + 1].lower() in ("rb", "br", "rf", "fr"):
                j -= 1
            i = j - 1
            continue

        if ch in (")", "]", "}"):
            depth += 1
        elif ch in ("(", "[", "{"):
            if depth == 0:
                # Found the unmatched opening paren
                j = i - 1
                while j >= 0 and code[j] in (" ", "\t"):
                    j -= 1
                name_end = j + 1
                while j >= 0 and (code[j].isalnum() or code[j] in ("_", ".")):
                    j -= 1
                name = code[j + 1:name_end].strip(".")
                if name and _IDENT_RE.fullmatch(name) and name not in candidates:
                    candidates.append(name)
                break
            depth -= 1
        i -= 1

    return candidates


def inspect_object(
    code: str,
    cursor_pos: int,
    session_id: Optional[str] = None,
) -> Dict[str, Any]:
    """Introspect the Python object at *cursor_pos* within *code*.

    Returns a dict with ``found``, ``name``, ``type``, ``signature``,
    ``docstring``, and ``detail`` fields.
    """
    empty = {
        "found": False,
        "name": "",
        "type": "",
        "signature": "",
        "docstring": "",
        "detail": "",
    }

    candidates = _extract_expression_at(code, cursor_pos)
    if not candidates:
        return empty

    session = _get_or_create_session_state(session_id)
    ns = session.namespace

    # Try each candidate until one resolves in the session namespace
    obj = None
    expr = ""
    for candidate in candidates:
        try:
            obj = eval(candidate, {"__builtins__": builtins.__dict__}, ns)  # noqa: S307
            expr = candidate
            break
        except Exception:
            continue

    if not expr:
        return empty

    result: Dict[str, Any] = {
        "found": True,
        "name": expr,
        "type": type(obj).__name__,
        "signature": "",
        "docstring": "",
        "detail": "",
    }

    # Signature (for callables)
    if callable(obj):
        try:
            sig = inspect.signature(obj)
            result["signature"] = f"{expr}{sig}"
        except (ValueError, TypeError):
            result["signature"] = f"{expr}(...)"

    # Docstring
    doc = inspect.getdoc(obj)
    if doc:
        if len(doc) > _MAX_DOC_LEN:
            doc = doc[:_MAX_DOC_LEN] + "\n..."
        result["docstring"] = doc

    # Extra detail for common types
    detail_parts: List[str] = []
    try:
        if isinstance(obj, dict):
            keys = list(obj.keys())
            shown = keys[:_MAX_DICT_KEYS]
            key_strs = [repr(k) for k in shown]
            detail_parts.append(f"Keys ({len(keys)}): [{', '.join(key_strs)}{'...' if len(keys) > _MAX_DICT_KEYS else ''}]")
            detail_parts.append(f"Length: {len(obj)}")
        elif isinstance(obj, pd.DataFrame):
            cols = list(obj.columns)
            shown = cols[:_MAX_DF_COLS]
            detail_parts.append(f"Shape: {obj.shape}")
            detail_parts.append(f"Columns ({len(cols)}): {shown}{'...' if len(cols) > _MAX_DF_COLS else ''}")
            dtypes_str = ", ".join(f"{c}: {obj[c].dtype}" for c in shown)
            detail_parts.append(f"Dtypes: {dtypes_str}")
        elif isinstance(obj, pd.Series):
            detail_parts.append(f"Shape: {obj.shape}")
            detail_parts.append(f"Dtype: {obj.dtype}")
            detail_parts.append(f"Name: {obj.name}")
        elif isinstance(obj, (list, tuple, set, frozenset)):
            detail_parts.append(f"Length: {len(obj)}")
            if len(obj) > 0:
                try:
                    preview = repr(obj)
                    if len(preview) > 200:
                        preview = preview[:200] + "..."
                    detail_parts.append(f"Preview: {preview}")
                except Exception:
                    pass
        elif isinstance(obj, str):
            detail_parts.append(f"Length: {len(obj)}")
            preview = repr(obj)
            if len(preview) > 200:
                preview = preview[:200] + "..."
            detail_parts.append(f"Value: {preview}")
        elif isinstance(obj, (int, float, complex, bool)):
            detail_parts.append(f"Value: {repr(obj)}")
        elif isinstance(obj, types.ModuleType):
            if hasattr(obj, "__version__"):
                detail_parts.append(f"Version: {obj.__version__}")
            if hasattr(obj, "__file__"):
                detail_parts.append(f"File: {obj.__file__}")
    except Exception:
        pass

    if detail_parts:
        result["detail"] = "\n".join(detail_parts)

    return result
