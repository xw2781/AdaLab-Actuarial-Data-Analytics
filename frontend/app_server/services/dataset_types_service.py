"""Dataset type definitions, normalization, and formula parsing."""
from __future__ import annotations

import os
import re
import json
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from fastapi import HTTPException

from app_server import config
from app_server.helpers import _canon_dataset_name, _parse_calculated_flag


def normalize_dataset_types_data(data: Any) -> Dict[str, Any]:
    rows_out: List[List[Any]] = []
    source_by_name: Dict[str, str] = {}

    if isinstance(data, list):
        for item in data:
            if not isinstance(item, dict):
                continue
            name = str(item.get("Name", "") or "")
            source = str(
                item.get("Source", item.get("source", item.get("SOURCE", ""))) or ""
            ).strip()
            if str(name).strip():
                source_by_name[str(name).strip()] = source
            rows_out.append([
                name,
                str(item.get("Data Format", "") or ""),
                str(item.get("Category", "") or ""),
                _parse_calculated_flag(item.get("Calculated", False)),
                str(item.get("Formula", "") or ""),
            ])
            if not rows_out[-1][3]:
                rows_out[-1][4] = ""
        return {
            "columns": list(config.DATASET_TYPES_COLUMNS),
            "rows": rows_out,
            "source_by_name": source_by_name,
        }

    if isinstance(data, dict):
        raw_cols = data.get("columns")
        raw_rows = data.get("rows")
        if isinstance(raw_cols, list) and isinstance(raw_rows, list):
            col_idx: Dict[str, int] = {}
            for i, c in enumerate(raw_cols):
                name = str(c or "").strip()
                if name:
                    col_idx[name] = i
                    col_idx[name.lower()] = i
            for r in raw_rows:
                if not isinstance(r, list):
                    continue
                i_name = col_idx.get("Name", -1)
                i_fmt = col_idx.get("Data Format", -1)
                i_cat = col_idx.get("Category", -1)
                i_calc = col_idx.get("Calculated", -1)
                i_formula = col_idx.get("Formula", -1)
                i_source = col_idx.get("Source", col_idx.get("source", -1))
                norm = [
                    str(r[i_name] if i_name >= 0 and i_name < len(r) and r[i_name] is not None else ""),
                    str(r[i_fmt] if i_fmt >= 0 and i_fmt < len(r) and r[i_fmt] is not None else ""),
                    str(r[i_cat] if i_cat >= 0 and i_cat < len(r) and r[i_cat] is not None else ""),
                    _parse_calculated_flag(r[i_calc] if i_calc >= 0 and i_calc < len(r) else False),
                    str(r[i_formula] if i_formula >= 0 and i_formula < len(r) and r[i_formula] is not None else ""),
                ]
                source = str(
                    r[i_source]
                    if i_source >= 0 and i_source < len(r) and r[i_source] is not None
                    else (r[5] if len(r) > 5 and r[5] is not None else "")
                ).strip()
                if not norm[3]:
                    norm[4] = ""
                rows_out.append(norm)
                if norm[0].strip():
                    source_by_name[norm[0].strip()] = source
            return {
                "columns": list(config.DATASET_TYPES_COLUMNS),
                "rows": rows_out,
                "source_by_name": source_by_name,
            }

    return {"columns": list(config.DATASET_TYPES_COLUMNS), "rows": [], "source_by_name": {}}


def _normalize_dataset_types_header_row(values: List[Any]) -> List[str]:
    out: List[str] = []
    for v in values:
        out.append(str(v if v is not None else "").strip())
    return out


def parse_local_dataset_types_file(file_path: str) -> Dict[str, Any]:
    path = str(file_path or "").strip()
    if not path:
        raise HTTPException(400, "file_path is required.")
    if not os.path.exists(path):
        raise HTTPException(404, f"File not found: {path}")

    ext = os.path.splitext(path)[1].strip().lower()
    if ext == ".json":
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
        except json.JSONDecodeError as e:
            raise HTTPException(400, f"Invalid JSON file: {str(e)}")
        except Exception as e:
            raise HTTPException(500, f"Failed to read JSON file: {str(e)}")
        normalized = normalize_dataset_types_data(raw)
        return {
            "format": "json",
            "columns": list(config.DATASET_TYPES_COLUMNS),
            "rows": list(normalized.get("rows") or []),
        }

    if ext != ".xlsx":
        raise HTTPException(400, f"Unsupported file extension: {ext}. Only .json and .xlsx are allowed.")

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid XLSX file: {str(e)}")

    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_values = next(rows_iter, None)
        if header_values is None:
            raise HTTPException(400, "XLSX file is empty.")

        header_raw = _normalize_dataset_types_header_row(list(header_values))
        while header_raw and header_raw[-1] == "":
            header_raw.pop()
        expected_columns_5 = list(config.DATASET_TYPES_COLUMNS)
        expected_columns_6 = ["Name", "Data Format", "Category", "Calculated", "Formula", "Source"]
        expected_columns_7 = list(config.DATASET_TYPES_FILE_COLUMNS)
        if header_raw == expected_columns_5:
            expected_columns = expected_columns_5
        elif header_raw == expected_columns_6:
            expected_columns = expected_columns_6
        elif header_raw == expected_columns_7:
            expected_columns = expected_columns_7
        else:
            raise HTTPException(
                400,
                "Invalid XLSX header. Expected either "
                f"[{', '.join(expected_columns_5)}], [{', '.join(expected_columns_6)}], "
                f"or [{', '.join(expected_columns_7)}].",
            )
        expected_count = len(expected_columns)
        idx_calculated = expected_columns.index("Calculated")
        idx_formula = expected_columns.index("Formula")

        parsed_rows: List[List[Any]] = []
        for row_values in rows_iter:
            row = list(row_values) if row_values is not None else []
            values = [row[i] if i < len(row) else "" for i in range(expected_count)]
            norm: List[Any] = []
            for i, value in enumerate(values):
                if i == idx_calculated:
                    norm.append(_parse_calculated_flag(value))
                else:
                    norm.append(str(value if value is not None else "").strip())
            if not bool(norm[idx_calculated]):
                norm[idx_formula] = ""

            row_name = str(norm[0] if len(norm) > 0 and norm[0] is not None else "").strip()
            row_fmt = str(norm[1] if len(norm) > 1 and norm[1] is not None else "").strip()
            row_cat = str(norm[2] if len(norm) > 2 and norm[2] is not None else "").strip()
            row_formula = str(norm[idx_formula] if idx_formula < len(norm) and norm[idx_formula] is not None else "").strip()
            row_calc = bool(norm[idx_calculated]) if idx_calculated < len(norm) else False
            if row_name != "" or row_fmt != "" or row_cat != "" or row_calc or row_formula != "":
                parsed_rows.append(norm)

        return {
            "format": "xlsx",
            "sheet_name": str(ws.title or "").strip(),
            "columns": expected_columns,
            "rows": parsed_rows,
        }
    finally:
        wb.close()


def _get_dataset_types_xlsx_path(json_path: str) -> str:
    base, _ext = os.path.splitext(str(json_path or "").strip())
    return base + ".xlsx"


def _xlsx_text_width(value: Any) -> int:
    if value is None:
        return 0
    if isinstance(value, bool):
        text = "TRUE" if value else "FALSE"
    else:
        text = str(value)
    if not text:
        return 0
    parts = text.splitlines() or [text]
    return max(len(part) for part in parts)


def _write_dataset_types_xlsx(tmp_xlsx_path: str, rows: List[List[Any]]) -> None:
    min_width = 8
    max_width = 96
    pad = 2
    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        ws.title = "Dataset Types"
        headers = list(config.DATASET_TYPES_FILE_COLUMNS)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        col_count = len(headers)
        col_widths = [_xlsx_text_width(h) for h in headers]
        for row in rows if isinstance(rows, list) else []:
            values = row if isinstance(row, list) else []
            out = [values[i] if i < len(values) else "" for i in range(col_count)]
            out[3] = _parse_calculated_flag(out[3] if len(out) > 3 else False)
            ws.append(out)
            for i, value in enumerate(out):
                w = _xlsx_text_width(value)
                if w > col_widths[i]:
                    col_widths[i] = w
        for i, width in enumerate(col_widths, start=1):
            final_width = max(min_width, min(max_width, width + pad))
            ws.column_dimensions[get_column_letter(i)].width = float(final_width)
        wb.save(tmp_xlsx_path)
    finally:
        wb.close()


def save_dataset_types_payload(filepath: str, payload: Dict[str, Any]) -> Dict[str, str]:
    json_path = str(filepath or "").strip()
    if not json_path:
        raise ValueError("filepath is required")

    xlsx_path = _get_dataset_types_xlsx_path(json_path)
    json_tmp_path = json_path + ".tmp"
    xlsx_tmp_path = xlsx_path + ".tmp"
    xlsx_rollback_tmp_path = xlsx_path + ".rollback.tmp"
    previous_xlsx_bytes: Optional[bytes] = None
    xlsx_replaced = False

    try:
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        if os.path.exists(xlsx_path):
            with open(xlsx_path, "rb") as f_prev_xlsx:
                previous_xlsx_bytes = f_prev_xlsx.read()

        rows = payload.get("rows", []) if isinstance(payload, dict) else []
        _write_dataset_types_xlsx(xlsx_tmp_path, rows if isinstance(rows, list) else [])
        os.replace(xlsx_tmp_path, xlsx_path)
        xlsx_replaced = True

        with open(json_tmp_path, "w", encoding="utf-8") as f_json:
            json.dump(payload, f_json, indent=2, ensure_ascii=False)
        os.replace(json_tmp_path, json_path)

        return {"json_path": json_path, "xlsx_path": xlsx_path}
    except Exception:
        if xlsx_replaced:
            try:
                if previous_xlsx_bytes is None:
                    if os.path.exists(xlsx_path):
                        os.remove(xlsx_path)
                else:
                    with open(xlsx_rollback_tmp_path, "wb") as f_rollback:
                        f_rollback.write(previous_xlsx_bytes)
                    os.replace(xlsx_rollback_tmp_path, xlsx_path)
            except Exception:
                pass
        raise
    finally:
        for p in (json_tmp_path, xlsx_tmp_path, xlsx_rollback_tmp_path):
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass


def get_dataset_type_names(project_name: str) -> List[str]:
    try:
        filepath = config.get_dataset_types_path(project_name)
    except ValueError:
        return []
    if not os.path.exists(filepath):
        return []
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            raw = json.load(f)
        data = normalize_dataset_types_data(raw)
        names: List[str] = []
        seen = set()
        for row in data.get("rows", []):
            if not isinstance(row, list) or not row:
                continue
            name = str(row[0] if row[0] is not None else "").strip()
            if not name or name in seen:
                continue
            seen.add(name)
            names.append(name)
        return names
    except Exception:
        return []


def _load_dataset_source_map(project_name: str) -> Dict[str, str]:
    try:
        filepath = config.get_field_mapping_path(project_name)
    except ValueError:
        return {}
    if not os.path.exists(filepath):
        return {}

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return {}

    rows = raw.get("rows", []) if isinstance(raw, dict) else []
    if not isinstance(rows, list):
        return {}

    out: Dict[str, str] = {}
    for row in rows:
        if not isinstance(row, dict):
            continue
        significance = str(row.get("significance", "") or "").strip()
        if significance != "Dataset":
            continue
        dataset_type = str(row.get("dataset_type", "") or "").strip()
        field_name = str(row.get("field_name", "") or "").strip()
        if not dataset_type or not field_name:
            continue
        key = _canon_dataset_name(dataset_type)
        if not key:
            continue
        if key not in out:
            out[key] = field_name
    return out


def _load_field_mapping_field_names(project_name: str) -> List[str]:
    try:
        filepath = config.get_field_mapping_path(project_name)
    except ValueError:
        return []
    if not os.path.exists(filepath):
        return []

    try:
        with open(filepath, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return []

    rows = raw.get("rows", []) if isinstance(raw, dict) else []
    if not isinstance(rows, list):
        return []

    out: List[str] = []
    seen: set[str] = set()
    for row in rows:
        if not isinstance(row, dict):
            continue
        field_name = str(row.get("field_name", "") or "").strip()
        if not field_name:
            continue
        key = re.sub(r"\s+", " ", field_name).casefold()
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(field_name)
    return out


def _is_source_generated_from_field_names(source: str, field_names: List[str]) -> bool:
    text = str(source or "").strip()
    if not text:
        return False
    candidates = [str(name or "").strip() for name in field_names if str(name or "").strip()]
    if not candidates:
        return False

    remaining = text
    matched_field = False
    for name in sorted(set(candidates), key=len, reverse=True):
        escaped = re.escape(name)
        if re.match(r"^[A-Za-z0-9_]+$", name):
            pattern = re.compile(rf"(?<![A-Za-z0-9_]){escaped}(?![A-Za-z0-9_])", flags=re.IGNORECASE)
        else:
            pattern = re.compile(escaped, flags=re.IGNORECASE)
        remaining, replacements = pattern.subn(" ", remaining)
        if replacements > 0:
            matched_field = True

    remaining = re.sub(r"(?:\d+(?:\.\d*)?|\.\d+)", " ", remaining)
    remaining = re.sub(r"[\s()+\-*/.,]+", " ", remaining)
    return matched_field and remaining.strip() == ""


def _wrap_source_component(expr: str) -> str:
    s = str(expr or "").strip()
    if not s:
        return ""
    if s.startswith("(") and s.endswith(")"):
        return s
    if re.search(r"\s[+\-*/]\s", s):
        return f"({s})"
    return s


def _replace_formula_components_with_sources(
    formula: str,
    resolve_component,
    known_dataset_names: List[str],
    source_map: Dict[str, str],
) -> str:
    text = str(formula or "").strip()
    if not text:
        return ""

    unique_names = sorted(
        set([str(n or "").strip() for n in known_dataset_names if str(n or "").strip()]),
        key=len,
        reverse=True,
    )
    placeholders: Dict[str, str] = {}

    def _resolve_component_value(token: str) -> str:
        resolved = str(resolve_component(token) or "").strip()
        if not resolved:
            resolved = str(source_map.get(_canon_dataset_name(token), "") or "").strip()
        if not resolved:
            resolved = str(token or "").strip()
        return _wrap_source_component(resolved)

    def _new_placeholder(value: str) -> str:
        key = f"__SRCREF_{len(placeholders)}__"
        placeholders[key] = value
        return key

    out = re.sub(
        r'"([^"]+)"',
        lambda m: _new_placeholder(_resolve_component_value(str(m.group(1) or "").strip())),
        text,
    )

    for name in unique_names:
        pattern = re.compile(
            rf"(?<![A-Za-z0-9_]){re.escape(name)}(?![A-Za-z0-9_])",
            flags=re.IGNORECASE,
        )
        out = pattern.sub(lambda _m, n=name: _new_placeholder(_resolve_component_value(n)), out)

    for key, value in placeholders.items():
        out = out.replace(key, value)

    out = out.replace('"', "")
    out = re.sub(r"\s+", " ", out).strip()
    return out


def _extract_formula_components(formula: str, known_dataset_names: List[str]) -> List[str]:
    text = str(formula or "").strip()
    if not text:
        return []

    quoted = [str(x or "").strip() for x in re.findall(r'"([^"]+)"', text)]
    quoted = [q for q in quoted if q]
    if quoted:
        seen: set[str] = set()
        out: List[str] = []
        for q in quoted:
            k = _canon_dataset_name(q)
            if not k or k in seen:
                continue
            seen.add(k)
            out.append(q)
        return out

    unique_names = sorted(
        set([str(n or "").strip() for n in known_dataset_names if str(n or "").strip()]),
        key=len,
        reverse=True,
    )
    if not unique_names:
        return []

    matches: List[Tuple[int, int, str]] = []
    for name in unique_names:
        pattern = re.compile(
            rf"(?<![A-Za-z0-9_]){re.escape(name)}(?![A-Za-z0-9_])",
            flags=re.IGNORECASE,
        )
        for m in pattern.finditer(text):
            matches.append((m.start(), m.end(), name))
    if not matches:
        return []

    matches.sort(key=lambda x: (x[0], -(x[1] - x[0])))
    used: List[Tuple[int, int]] = []
    out_list: List[str] = []
    seen_keys: set[str] = set()
    for start, end, name in matches:
        overlap = False
        for us, ue in used:
            if start < ue and end > us:
                overlap = True
                break
        if overlap:
            continue
        used.append((start, end))
        k = _canon_dataset_name(name)
        if not k or k in seen_keys:
            continue
        seen_keys.add(k)
        out_list.append(name)
    return out_list


def _find_unresolved_dataset_refs(text: str, known_dataset_names: List[str]) -> List[str]:
    s = str(text or "").strip()
    if not s:
        return []
    unresolved: List[str] = []
    seen: set[str] = set()
    for name in sorted(
        set([str(n or "").strip() for n in known_dataset_names if str(n or "").strip()]),
        key=len,
        reverse=True,
    ):
        pattern = re.compile(
            rf"(?<![A-Za-z0-9_]){re.escape(name)}(?![A-Za-z0-9_])",
            flags=re.IGNORECASE,
        )
        if pattern.search(s):
            k = _canon_dataset_name(name)
            if k and k not in seen:
                seen.add(k)
                unresolved.append(name)
    return unresolved


def _build_dataset_source_resolver(
    dataset_rows: List[List[Any]],
    source_map: Dict[str, str],
):
    formula_by_key: Dict[str, str] = {}
    calculated_by_key: Dict[str, bool] = {}
    names_by_key: Dict[str, str] = {}
    for row in dataset_rows:
        if not isinstance(row, list):
            continue
        name = str(row[0] if len(row) > 0 and row[0] is not None else "").strip()
        calculated = _parse_calculated_flag(row[3] if len(row) > 3 else False)
        formula = str(row[4] if len(row) > 4 and row[4] is not None else "").strip()
        key = _canon_dataset_name(name)
        if not key:
            continue
        names_by_key.setdefault(key, name)
        calculated_by_key[key] = calculated
        formula_by_key[key] = formula

    memo: Dict[str, str] = {}
    known_dataset_names = list(names_by_key.values())

    def resolve_dataset_source(dataset_name: str, stack: Optional[set[str]] = None) -> str:
        key = _canon_dataset_name(dataset_name)
        if not key:
            return ""
        if key in memo:
            return memo[key]
        if stack is None:
            stack = set()
        if key in stack:
            return source_map.get(key, "")

        stack2 = set(stack)
        stack2.add(key)
        base = source_map.get(key, "")
        is_calculated = calculated_by_key.get(key, False)
        formula = formula_by_key.get(key, "").strip()
        if (not is_calculated) or formula == "":
            memo[key] = base
            return memo[key]

        expr = _replace_formula_components_with_sources(
            formula=formula,
            resolve_component=lambda comp: resolve_dataset_source(comp, stack2),
            known_dataset_names=known_dataset_names,
            source_map=source_map,
        )
        memo[key] = expr if expr else base
        return memo[key]

    return resolve_dataset_source
