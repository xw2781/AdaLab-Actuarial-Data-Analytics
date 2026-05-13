"""Reserving class business logic: types, values, combinations, path tree, hidden paths, filter specs."""
from __future__ import annotations

import os
import re
import json
import hashlib
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from fastapi import HTTPException

from app_server import config
from app_server.config import (
    RESERVING_CLASS_TYPES_SHEET_NAME,
    RESERVING_CLASS_TYPES_COLUMNS,
    RESERVING_CLASS_TYPES_FILE_COLUMNS,
    RESERVING_CLASS_PATH_TREE_MAX_GENERATED,
    _RESERVING_CLASS_PATH_TREE_LOCK,
    _RESERVING_CLASS_FILTER_SPEC_LOCK,
    get_reserving_class_types_path,
    get_reserving_class_values_path,
    get_reserving_class_combinations_path,
    get_reserving_class_path_tree_path,
    get_reserving_class_filter_spec_pref_path,
    get_project_settings_workbook_path,
    get_cache_path,
    get_field_mapping_path,
)
from app_server.helpers import (
    _norm_tree_path,
    _canon_reserving_class_type_name,
    _normalize_reserving_hidden_path_list,
    _normalize_reserving_filter_spec,
    _normalize_reserving_filter_preferences,
    _format_excel_cell_as_text,
    _normalize_sheet_columns,
    _normalize_sheet_rows,
    _normalize_formula_operator_spacing,
    _canon_project_pref_key,
    _canon_dataset_name,
    _parse_positive_int,
)
from app_server.services.dataset_types_service import (
    _replace_formula_components_with_sources,
    _extract_formula_components,
)
from app_server.services.project_user_preferences_service import (
    get_preferences as get_project_user_preferences,
    update_preferences as update_project_user_preferences,
)

PROJECT_USER_RESERVING_CLASS_TREE_KEY = "reservingClassTree"


# ---------------------------------------------------------------------------
# Project-user reserving class tree preferences
# ---------------------------------------------------------------------------

def _project_user_preference_data(project_name: str) -> Tuple[Dict[str, Any], str]:
    out = get_project_user_preferences(project_name)
    data = out.get("data", {}) if isinstance(out, dict) else {}
    path = str(out.get("path", "") or "") if isinstance(out, dict) else ""
    return (data if isinstance(data, dict) else {}), path


def _project_user_reserving_tree_section(project_name: str) -> Tuple[Dict[str, Any], str]:
    data, path = _project_user_preference_data(project_name)
    section = data.get(PROJECT_USER_RESERVING_CLASS_TREE_KEY, {})
    return (section if isinstance(section, dict) else {}), path


def _read_reserving_class_tree_preferences(project_name: str) -> Dict[str, Any]:
    section, _path = _project_user_reserving_tree_section(project_name)
    return _normalize_reserving_filter_preferences(section.get("preferences", {}))


def _write_reserving_class_tree_preferences(project_name: str, preferences: Any) -> Dict[str, Any]:
    preferences_norm = _normalize_reserving_filter_preferences(preferences)
    update_project_user_preferences(project_name, {
        PROJECT_USER_RESERVING_CLASS_TREE_KEY: {
            "preferences": preferences_norm,
            "updated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        },
    })
    return preferences_norm

def get_hidden_paths_for_project(project_name: str) -> Dict[str, Any]:
    key = _canon_project_pref_key(project_name)
    if not key:
        raise ValueError("project_name is required")
    section, filepath = _project_user_reserving_tree_section(project_name)
    hidden_paths = _normalize_reserving_hidden_path_list(
        section.get("hiddenPaths", []) if isinstance(section, dict) else []
    )
    return {
        "path": filepath,
        "project_key": key,
        "hidden_paths": hidden_paths,
    }

def save_hidden_paths_for_project(project_name: str, hidden_paths: Any) -> Dict[str, Any]:
    key = _canon_project_pref_key(project_name)
    if not key:
        raise ValueError("project_name is required")

    hidden_paths_norm = _normalize_reserving_hidden_path_list(hidden_paths)
    now_iso = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    out = update_project_user_preferences(project_name, {
        PROJECT_USER_RESERVING_CLASS_TREE_KEY: {
            "hiddenPaths": hidden_paths_norm,
            "updated_at": now_iso,
        },
    })

    return {
        "path": str(out.get("path", "") or "") if isinstance(out, dict) else "",
        "project_key": key,
        "hidden_paths": hidden_paths_norm,
    }


# ---------------------------------------------------------------------------
# Filter spec store
# ---------------------------------------------------------------------------

def _default_reserving_filter_spec_store() -> Dict[str, Any]:
    return {
        "projects": {},
    }

def _normalize_reserving_filter_spec_store_payload(raw_store: Any) -> Dict[str, Any]:
    default_store = _default_reserving_filter_spec_store()
    if not isinstance(raw_store, dict):
        return default_store

    projects_raw = raw_store.get("projects", {})
    if not isinstance(projects_raw, dict):
        projects_raw = {}
    projects_norm: Dict[str, Dict[str, Any]] = {}
    for project_key_raw, entry_raw in projects_raw.items():
        project_key = str(project_key_raw or "").strip().lower()
        if not project_key:
            continue
        entry = entry_raw if isinstance(entry_raw, dict) else {}
        filter_spec = _normalize_reserving_filter_spec(entry.get("filter_spec", {}))
        projects_norm[project_key] = {
            "project_name": str(entry.get("project_name", "") or "").strip(),
            "updated_at": str(entry.get("updated_at", "") or "").strip(),
            "filter_spec": filter_spec,
        }

    return {
        "updated_at": str(raw_store.get("updated_at", "") or "").strip(),
        "projects": projects_norm,
    }

def _load_reserving_filter_spec_store(filepath: str) -> Dict[str, Any]:
    default_store = _default_reserving_filter_spec_store()
    if not os.path.exists(filepath):
        return default_store
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return default_store
    return _normalize_reserving_filter_spec_store_payload(raw)

def _load_and_cleanup_reserving_filter_spec_store(filepath: str) -> Dict[str, Any]:
    default_store = _default_reserving_filter_spec_store()
    if not os.path.exists(filepath):
        return default_store
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return default_store

    normalized = _normalize_reserving_filter_spec_store_payload(raw)
    raw_cmp = raw if isinstance(raw, dict) else {}
    if json.dumps(raw_cmp, sort_keys=True, ensure_ascii=False) != json.dumps(normalized, sort_keys=True, ensure_ascii=False):
        _write_reserving_filter_spec_store(filepath, normalized)
    return normalized

def _write_reserving_filter_spec_store(filepath: str, payload: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    tmp_path = filepath + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    os.replace(tmp_path, filepath)

def get_filter_spec_for_project(project_name: str) -> Dict[str, Any]:
    key = _canon_project_pref_key(project_name)
    if not key:
        raise ValueError("project_name is required")

    filepath = get_reserving_class_filter_spec_pref_path()
    with _RESERVING_CLASS_FILTER_SPEC_LOCK:
        store = _load_and_cleanup_reserving_filter_spec_store(filepath)
    projects = store.get("projects", {}) if isinstance(store, dict) else {}
    entry = projects.get(key, {}) if isinstance(projects, dict) else {}
    filter_spec = _normalize_reserving_filter_spec(
        entry.get("filter_spec", {}) if isinstance(entry, dict) else {}
    )
    preferences = _read_reserving_class_tree_preferences(project_name)
    return {
        "path": filepath,
        "project_key": key,
        "filter_spec": filter_spec,
        "preferences": preferences,
    }

def save_filter_spec_for_project(
    project_name: str,
    filter_spec: Any,
    preferences: Any = None,
) -> Dict[str, Any]:
    key = _canon_project_pref_key(project_name)
    if not key:
        raise ValueError("project_name is required")

    filepath = get_reserving_class_filter_spec_pref_path()
    filter_spec_norm = _normalize_reserving_filter_spec(filter_spec)
    now_iso = datetime.utcnow().isoformat(timespec="seconds") + "Z"
    preferences_out = (
        _write_reserving_class_tree_preferences(project_name, preferences)
        if preferences is not None
        else _read_reserving_class_tree_preferences(project_name)
    )

    with _RESERVING_CLASS_FILTER_SPEC_LOCK:
        store = _load_and_cleanup_reserving_filter_spec_store(filepath)
        projects = store.get("projects", {})
        if not isinstance(projects, dict):
            projects = {}

        if filter_spec_norm:
            projects[key] = {
                "project_name": str(project_name or "").strip(),
                "updated_at": now_iso,
                "filter_spec": filter_spec_norm,
                "preferences": {},
            }
        else:
            projects.pop(key, None)

        store["projects"] = projects
        store["updated_at"] = now_iso
        _write_reserving_filter_spec_store(filepath, store)

    return {
        "path": filepath,
        "project_key": key,
        "filter_spec": filter_spec_norm,
        "preferences": preferences_out,
    }


# ---------------------------------------------------------------------------
# Types management
# ---------------------------------------------------------------------------

def _read_reserving_class_types_sheet(path: str) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    if RESERVING_CLASS_TYPES_SHEET_NAME not in wb.sheetnames:
        return {"columns": list(RESERVING_CLASS_TYPES_COLUMNS), "rows": []}

    ws = wb[RESERVING_CLASS_TYPES_SHEET_NAME]
    max_row = int(ws.max_row or 1)
    max_col = int(ws.max_column or 1)

    raw_columns: List[str] = []
    for c in range(1, max_col + 1):
        raw_columns.append(_format_excel_cell_as_text(ws.cell(row=1, column=c).value))
    columns = _normalize_sheet_columns(raw_columns, list(RESERVING_CLASS_TYPES_COLUMNS))
    width = len(columns)

    rows: List[List[str]] = []
    for r in range(2, max_row + 1):
        row: List[str] = []
        for c in range(1, width + 1):
            row.append(_format_excel_cell_as_text(ws.cell(row=r, column=c).value).strip())
        if any(cell != "" for cell in row):
            rows.append(row)
    return {"columns": columns, "rows": rows}

def _write_reserving_class_types_sheet(path: str, columns: List[str], rows: List[List[str]]) -> Dict[str, Any]:
    columns_norm = _normalize_sheet_columns(columns, list(RESERVING_CLASS_TYPES_COLUMNS))
    rows_norm = _normalize_sheet_rows(rows, len(columns_norm))

    if os.path.exists(path):
        wb = openpyxl.load_workbook(path, data_only=False, keep_vba=True)
    else:
        wb = openpyxl.Workbook()
        # Use an explicit title for the first sheet when creating a new workbook.
        wb.active.title = RESERVING_CLASS_TYPES_SHEET_NAME

    ws = wb[RESERVING_CLASS_TYPES_SHEET_NAME] if RESERVING_CLASS_TYPES_SHEET_NAME in wb.sheetnames else wb.create_sheet(RESERVING_CLASS_TYPES_SHEET_NAME)
    if ws.max_row > 0:
        ws.delete_rows(1, ws.max_row)

    for idx, name in enumerate(columns_norm, start=1):
        ws.cell(row=1, column=idx, value=name)
    for r_idx, row in enumerate(rows_norm, start=2):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    os.makedirs(os.path.dirname(path), exist_ok=True)
    tmp_path = path + ".tmp"
    wb.save(tmp_path)
    os.replace(tmp_path, path)

    return {
        "columns": columns_norm,
        "rows": rows_norm,
    }

def _get_reserving_class_types_xlsx_path(json_path: str) -> str:
    base, _ext = os.path.splitext(str(json_path or "").strip())
    return base + ".xlsx"

def _xlsx_text_width(value: Any) -> int:
    if value is None:
        return 0
    text = str(value)
    if not text:
        return 0
    parts = text.splitlines() or [text]
    return max(len(part) for part in parts)

def _write_reserving_class_types_xlsx(tmp_xlsx_path: str, rows: List[List[Any]]) -> None:
    min_width = 8
    max_width = 96
    pad = 2
    wb = openpyxl.Workbook()
    try:
        ws = wb.active
        ws.title = "Reserving Class Types"
        headers = list(RESERVING_CLASS_TYPES_FILE_COLUMNS)
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)

        col_count = len(headers)
        col_widths = [_xlsx_text_width(h) for h in headers]
        for row in rows if isinstance(rows, list) else []:
            values = row if isinstance(row, list) else []
            out = [
                str(values[i] if i < len(values) and values[i] is not None else "").strip()
                for i in range(col_count)
            ]
            ws.append(out)
            for i, value in enumerate(out):
                w = _xlsx_text_width(value)
                if w > col_widths[i]:
                    col_widths[i] = w

        for i, width in enumerate(col_widths, start=1):
            final_width = max(min_width, min(max_width, width + pad))
            ws.column_dimensions[get_column_letter(i)].width = float(final_width)
        wb.save(tmp_xlsx_path)
    finally:
        wb.close()

def save_reserving_class_types_payload(filepath: str, payload: Dict[str, Any]) -> Dict[str, str]:
    json_path = str(filepath or "").strip()
    if not json_path:
        raise ValueError("filepath is required")

    xlsx_path = _get_reserving_class_types_xlsx_path(json_path)
    json_tmp_path = json_path + ".tmp"
    xlsx_tmp_path = xlsx_path + ".tmp"
    xlsx_rollback_tmp_path = xlsx_path + ".rollback.tmp"
    previous_xlsx_bytes: Optional[bytes] = None
    xlsx_replaced = False

    try:
        os.makedirs(os.path.dirname(json_path), exist_ok=True)
        if os.path.exists(xlsx_path):
            with open(xlsx_path, "rb") as f_prev_xlsx:
                previous_xlsx_bytes = f_prev_xlsx.read()

        rows = payload.get("rows", []) if isinstance(payload, dict) else []
        _write_reserving_class_types_xlsx(xlsx_tmp_path, rows if isinstance(rows, list) else [])
        os.replace(xlsx_tmp_path, xlsx_path)
        xlsx_replaced = True

        with open(json_tmp_path, "w", encoding="utf-8") as f_json:
            json.dump(payload, f_json, indent=2, ensure_ascii=False)
        os.replace(json_tmp_path, json_path)

        return {"json_path": json_path, "xlsx_path": xlsx_path}
    except Exception:
        if xlsx_replaced:
            try:
                if previous_xlsx_bytes is None:
                    if os.path.exists(xlsx_path):
                        os.remove(xlsx_path)
                else:
                    with open(xlsx_rollback_tmp_path, "wb") as f_rollback:
                        f_rollback.write(previous_xlsx_bytes)
                    os.replace(xlsx_rollback_tmp_path, xlsx_path)
            except Exception:
                pass
        raise
    finally:
        for p in (json_tmp_path, xlsx_tmp_path, xlsx_rollback_tmp_path):
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass

def normalize_reserving_class_types_data(data: Any) -> Dict[str, Any]:
    rows_out: List[List[str]] = []
    cols = list(RESERVING_CLASS_TYPES_FILE_COLUMNS)

    if isinstance(data, dict):
        raw_cols = data.get("columns")
        raw_rows = data.get("rows")
        col_idx: Dict[str, int] = {}
        if isinstance(raw_cols, list):
            for i, c in enumerate(raw_cols):
                name = str(c if c is not None else "").strip()
                if name:
                    col_idx[name] = i

        if isinstance(raw_rows, list):
            for raw in raw_rows:
                if not isinstance(raw, list):
                    continue
                i_name = col_idx.get("Name", 0 if len(raw) > 0 else -1)
                i_level = col_idx.get("Level", 1 if len(raw) > 1 else -1)
                i_formula = col_idx.get("Formula", 2 if len(raw) > 2 else -1)
                i_eex = col_idx.get("EEX Formula", 3 if len(raw) > 3 else -1)
                i_source = col_idx.get("Source", 4 if len(raw) > 4 else -1)
                row = [
                    str(raw[i_name] if i_name >= 0 and i_name < len(raw) and raw[i_name] is not None else "").strip(),
                    str(raw[i_level] if i_level >= 0 and i_level < len(raw) and raw[i_level] is not None else "").strip(),
                    str(raw[i_formula] if i_formula >= 0 and i_formula < len(raw) and raw[i_formula] is not None else "").strip(),
                    str(raw[i_eex] if i_eex >= 0 and i_eex < len(raw) and raw[i_eex] is not None else "").strip(),
                    str(raw[i_source] if i_source >= 0 and i_source < len(raw) and raw[i_source] is not None else "").strip(),
                ]
                if any(cell != "" for cell in row):
                    rows_out.append(row)

    return {"columns": cols, "rows": rows_out}

def parse_local_reserving_class_types_file(file_path: str) -> Dict[str, Any]:
    path = str(file_path or "").strip()
    if not path:
        raise HTTPException(400, "file_path is required.")
    if not os.path.exists(path):
        raise HTTPException(404, f"File not found: {path}")

    ext = os.path.splitext(path)[1].strip().lower()
    if ext == ".json":
        try:
            with open(path, "r", encoding="utf-8") as f:
                raw = json.load(f)
        except json.JSONDecodeError as e:
            raise HTTPException(400, f"Invalid JSON file: {str(e)}")
        except Exception as e:
            raise HTTPException(500, f"Failed to read JSON file: {str(e)}")
        normalized = normalize_reserving_class_types_data(raw)
        return {
            "format": "json",
            "columns": list(normalized.get("columns") or []),
            "rows": list(normalized.get("rows") or []),
        }

    if ext != ".xlsx":
        raise HTTPException(400, f"Unsupported file extension: {ext}. Only .json and .xlsx are allowed.")

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Invalid XLSX file: {str(e)}")

    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header_values = next(rows_iter, None)
        if header_values is None:
            raise HTTPException(400, "XLSX file is empty.")

        header_raw = [str(v if v is not None else "").strip() for v in list(header_values)]
        while header_raw and header_raw[-1] == "":
            header_raw.pop()

        expected_columns_4 = list(RESERVING_CLASS_TYPES_COLUMNS)
        expected_columns_5 = list(RESERVING_CLASS_TYPES_FILE_COLUMNS)
        if header_raw == expected_columns_4:
            expected_columns = expected_columns_4
        elif header_raw == expected_columns_5:
            expected_columns = expected_columns_5
        else:
            raise HTTPException(
                400,
                "Invalid XLSX header. Expected either "
                f"[{', '.join(expected_columns_4)}] or [{', '.join(expected_columns_5)}].",
            )

        expected_count = len(expected_columns)
        idx_formula = expected_columns.index("Formula")
        idx_eex_formula = expected_columns.index("EEX Formula")
        parsed_rows: List[List[str]] = []
        for row_values in rows_iter:
            row = list(row_values) if row_values is not None else []
            values = [row[i] if i < len(row) else "" for i in range(expected_count)]
            norm = [str(v if v is not None else "").strip() for v in values]
            norm[idx_formula] = _normalize_formula_operator_spacing(norm[idx_formula])
            norm[idx_eex_formula] = _normalize_formula_operator_spacing(norm[idx_eex_formula])
            if any(cell != "" for cell in norm):
                parsed_rows.append(norm)

        return {
            "format": "xlsx",
            "sheet_name": str(ws.title or "").strip(),
            "columns": expected_columns,
            "rows": parsed_rows,
        }
    finally:
        wb.close()

def _to_reserving_class_types_ui_data(data: Dict[str, Any]) -> Dict[str, Any]:
    rows_out: List[List[str]] = []
    rows = data.get("rows", []) if isinstance(data, dict) else []
    if isinstance(rows, list):
        for raw in rows:
            if not isinstance(raw, list):
                continue
            row = [
                str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip(),
                str(raw[1] if len(raw) > 1 and raw[1] is not None else "").strip(),
                str(raw[2] if len(raw) > 2 and raw[2] is not None else "").strip(),
                str(raw[3] if len(raw) > 3 and raw[3] is not None else "").strip(),
            ]
            if any(cell != "" for cell in row):
                rows_out.append(row)
    return {"columns": list(RESERVING_CLASS_TYPES_COLUMNS), "rows": rows_out}

def _load_reserving_class_types_raw_data(project_name: str) -> Dict[str, Any]:
    json_path = get_reserving_class_types_path(project_name)
    if os.path.exists(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                return normalize_reserving_class_types_data(json.load(f))
        except Exception:
            pass

    # One-way migration bootstrap: seed from settings.xlsx sheet if JSON does not exist yet.
    xlsx_path = get_project_settings_workbook_path(project_name)
    if os.path.exists(xlsx_path):
        try:
            xlsx_data = _read_reserving_class_types_sheet(xlsx_path)
            return normalize_reserving_class_types_data(xlsx_data)
        except Exception:
            pass

    return {"columns": list(RESERVING_CLASS_TYPES_FILE_COLUMNS), "rows": []}

def _load_reserving_class_value_fields(project_name: str) -> List[Dict[str, Any]]:
    try:
        path = get_reserving_class_values_path(project_name)
    except ValueError:
        return []
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return []
    fields = raw.get("fields", []) if isinstance(raw, dict) else []
    out: List[Dict[str, Any]] = []
    if isinstance(fields, list):
        for item in fields:
            if isinstance(item, dict):
                out.append(item)
    return out

def _build_reserving_class_source_rows(
    source_fields: List[Dict[str, Any]],
) -> Tuple[List[List[str]], set[str], set[Tuple[str, str]]]:
    source_name_level_pairs: set[Tuple[str, str]] = set()
    display_name_by_pair: Dict[Tuple[str, str], str] = {}
    for field in source_fields:
        if not isinstance(field, dict):
            continue
        level_val = field.get("level")
        level: Optional[int] = None
        try:
            if level_val is not None and str(level_val).strip() != "":
                n = int(level_val)
                if n >= 1:
                    level = n
        except Exception:
            level = None

        distinct_values = field.get("distinct_values", [])
        if not isinstance(distinct_values, list):
            continue
        for raw in distinct_values:
            name = str(raw if raw is not None else "").strip()
            if not name:
                continue
            name_key = _canon_reserving_class_type_name(name)
            if not name_key:
                continue
            level_text = str(level) if isinstance(level, int) and level >= 1 else ""
            pair = (name_key, level_text)
            source_name_level_pairs.add(pair)
            # Keep first-seen display casing for each (name, level) pair.
            display_name_by_pair.setdefault(pair, name)

    sortable: List[Tuple[int, str, str, str]] = []
    for (name_key, level_text), name in display_name_by_pair.items():
        lvl_sort = int(level_text) if level_text.isdigit() else 10**9
        sortable.append((lvl_sort, name.lower(), name, level_text))
    sortable.sort()

    rows: List[List[str]] = []
    source_names: set[str] = set()
    for _lvl_sort, _name_sort, name, level_text in sortable:
        source_names.add(_canon_reserving_class_type_name(name))
        rows.append([name, level_text, "", ""])
    return (rows, source_names, source_name_level_pairs)

def _merge_reserving_class_types_rows(
    existing_rows: List[List[str]],
    source_rows: List[List[str]],
    source_name_level_pairs: set[Tuple[str, str]],
) -> List[List[str]]:
    row_by_key: Dict[str, List[str]] = {}
    key_order: List[str] = []

    def _pair_key(name: str, level: str) -> Tuple[str, str]:
        return (_canon_reserving_class_type_name(name), str(level or "").strip())

    def _row_key(name: str, level: str) -> str:
        pair = _pair_key(name, level)
        # Source-derived rows are keyed by (name, level) so same name can exist at multiple levels.
        if pair in source_name_level_pairs:
            return f"{pair[0]}|{pair[1]}"
        # User-defined rows keep historical behavior: unique by name.
        return pair[0]

    # Seed with source-derived rows first (baseline).
    for raw in source_rows:
        if not isinstance(raw, list):
            continue
        name = str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip()
        if not name:
            continue
        level = str(raw[1] if len(raw) > 1 and raw[1] is not None else "").strip()
        formula = str(raw[2] if len(raw) > 2 and raw[2] is not None else "").strip()
        eex_formula = str(raw[3] if len(raw) > 3 and raw[3] is not None else "").strip()
        key = _row_key(name, level)
        if key not in row_by_key:
            key_order.append(key)
        row_by_key[key] = [name, level, formula, eex_formula]

    # User-defined rows (formula non-empty) always win on conflicts.
    for raw in existing_rows:
        if not isinstance(raw, list):
            continue
        name = str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip()
        level = str(raw[1] if len(raw) > 1 and raw[1] is not None else "").strip()
        formula = str(raw[2] if len(raw) > 2 and raw[2] is not None else "").strip()
        eex_formula = str(raw[3] if len(raw) > 3 and raw[3] is not None else "").strip()
        if not name:
            continue
        # User-defined rows must have Formula populated.
        if not formula:
            continue
        key = _row_key(name, level)
        if key not in row_by_key:
            key_order.append(key)
        row_by_key[key] = [name, level, formula, eex_formula]

    merged: List[List[str]] = []
    for key in key_order:
        row = row_by_key.get(key)
        if row is None:
            continue
        merged.append(row)
    return merged

def _reserving_class_row_key(name: str, level: str, source_name_level_pairs: set[Tuple[str, str]]) -> str:
    pair = (_canon_reserving_class_type_name(name), str(level or "").strip())
    if pair in source_name_level_pairs:
        return f"{pair[0]}|{pair[1]}"
    return pair[0]

def _get_changed_reserving_class_formula_rows(
    previous_rows: List[List[str]],
    next_rows: List[List[str]],
    source_name_level_pairs: set[Tuple[str, str]],
) -> List[List[str]]:
    previous_by_key: Dict[str, List[str]] = {}
    for raw in previous_rows:
        if not isinstance(raw, list):
            continue
        name = str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip()
        level = str(raw[1] if len(raw) > 1 and raw[1] is not None else "").strip()
        key = _reserving_class_row_key(name, level, source_name_level_pairs)
        if key:
            previous_by_key[key] = [
                name,
                level,
                str(raw[2] if len(raw) > 2 and raw[2] is not None else "").strip(),
                str(raw[3] if len(raw) > 3 and raw[3] is not None else "").strip(),
            ]

    changed_rows: List[List[str]] = []
    for raw in next_rows:
        if not isinstance(raw, list):
            continue
        name = str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip()
        level = str(raw[1] if len(raw) > 1 and raw[1] is not None else "").strip()
        formula = str(raw[2] if len(raw) > 2 and raw[2] is not None else "").strip()
        eex_formula = str(raw[3] if len(raw) > 3 and raw[3] is not None else "").strip()
        key = _reserving_class_row_key(name, level, source_name_level_pairs)
        if not key:
            continue
        previous = previous_by_key.get(key)
        if previous is None or previous != [name, level, formula, eex_formula]:
            changed_rows.append([name, level, formula, eex_formula])

    return changed_rows

def _extract_reserving_formula_component_refs(formula: str, known_names: List[str]) -> List[Tuple[str, bool]]:
    text = str(formula or "").strip()
    if not text:
        return []

    out_list: List[Tuple[str, bool]] = []
    seen_quoted: set[str] = set()
    used: List[Tuple[int, int]] = []
    for match in re.finditer(r'"([^"]*)"', text):
        value = str(match.group(1) if match.group(1) is not None else "")
        used.append((match.start(), match.end()))
        if not value or value in seen_quoted:
            continue
        seen_quoted.add(value)
        out_list.append((value, True))

    unique_names = sorted(
        set([str(n or "").strip() for n in known_names if str(n or "").strip()]),
        key=len,
        reverse=True,
    )
    matches: List[Tuple[int, int, str]] = []
    for name in unique_names:
        pattern = re.compile(
            rf"(?<![A-Za-z0-9_]){re.escape(name)}(?![A-Za-z0-9_])",
            flags=re.IGNORECASE,
        )
        for match in pattern.finditer(text):
            matches.append((match.start(), match.end(), name))

    matches.sort(key=lambda item: (item[0], -(item[1] - item[0])))
    seen_keys: set[str] = set()
    residual = text
    if used or matches:
        chars = list(text)
        for start, end in used:
            for idx in range(start, min(end, len(chars))):
                chars[idx] = " "
        for start, end, name in matches:
            overlap = False
            for used_start, used_end in used:
                if start < used_end and end > used_start:
                    overlap = True
                    break
            if overlap:
                continue
            used.append((start, end))
            key = _canon_reserving_class_type_name(name)
            if not key or key in seen_keys:
                continue
            seen_keys.add(key)
            out_list.append((name, False))
            for idx in range(start, min(end, len(chars))):
                chars[idx] = " "
        residual = "".join(chars)

    token_parts = [
        part.strip()
        for part in re.split(r"[+\-*/]", re.sub(r"[()]", " ", residual))
        if str(part or "").strip()
    ]
    for token in token_parts:
        if re.fullmatch(r"\d+(\.\d+)?", token):
            continue
        key = _canon_reserving_class_type_name(token)
        if not key or key in seen_keys:
            continue
        seen_keys.add(key)
        out_list.append((token, False))

    return out_list


def _reserving_formula_name_requires_quotes(name: str) -> bool:
    return bool(re.search(r"[+\-*/]", str(name or "")))

def _validate_reserving_class_formula_components(
    rows: List[List[str]],
    source_rows: List[List[str]],
    known_rows: Optional[List[List[str]]] = None,
) -> None:
    known_names: List[str] = []
    for raw in source_rows:
        if isinstance(raw, list):
            name = str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip()
            if name:
                known_names.append(name)
    for raw in (known_rows if known_rows is not None else rows):
        if isinstance(raw, list):
            name = str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip()
            if name:
                known_names.append(name)
    known_exact_names = {name for name in known_names if name}
    known_keys = {
        _canon_reserving_class_type_name(name)
        for name in known_names
        if _canon_reserving_class_type_name(name)
    }

    issues: List[str] = []
    for raw in rows:
        if not isinstance(raw, list):
            continue
        row_name = str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip() or "(blank name)"
        formula_fields = [
            ("Formula", str(raw[2] if len(raw) > 2 and raw[2] is not None else "").strip()),
            ("EEX Formula", str(raw[3] if len(raw) > 3 and raw[3] is not None else "").strip()),
        ]
        for field_label, formula in formula_fields:
            if not formula:
                continue
            components = _extract_reserving_formula_component_refs(formula, known_names)
            invalid = []
            seen_invalid_exact: set[str] = set()
            seen_invalid_canon: set[str] = set()
            for component, is_quoted in components:
                if is_quoted:
                    if not component or component in known_exact_names or component in seen_invalid_exact:
                        continue
                    seen_invalid_exact.add(component)
                    invalid.append(component)
                    continue
                if _reserving_formula_name_requires_quotes(component):
                    if component not in seen_invalid_exact:
                        seen_invalid_exact.add(component)
                        invalid.append(component)
                    continue
                key = _canon_reserving_class_type_name(component)
                if not key or key in known_keys or key in seen_invalid_canon:
                    continue
                seen_invalid_canon.add(key)
                invalid.append(component)
            if invalid:
                issues.append(f"{row_name} - {field_label}: {', '.join(invalid)}")

    if issues:
        raise ValueError(
            "Reserving class formulas can only reference existing reserving class type names, and names containing +, -, *, or / must be quoted. "
            + "; ".join(issues)
        )

def _build_reserving_class_type_source_resolver(
    reserving_rows: List[List[str]],
    source_name_map: Dict[str, str],
):
    formula_by_key: Dict[str, str] = {}
    names_by_key: Dict[str, str] = {}
    for row in reserving_rows:
        if not isinstance(row, list):
            continue
        name = str(row[0] if len(row) > 0 and row[0] is not None else "").strip()
        formula = str(row[2] if len(row) > 2 and row[2] is not None else "").strip()
        key = _canon_reserving_class_type_name(name)
        if not key:
            continue
        names_by_key.setdefault(key, name)
        formula_by_key[key] = formula

    memo: Dict[str, str] = {}
    known_names = list(names_by_key.values())

    def resolve_reserving_source(type_name: str, stack: Optional[set[str]] = None) -> str:
        key = _canon_reserving_class_type_name(type_name)
        if not key:
            return ""
        if key in memo:
            return memo[key]
        if stack is None:
            stack = set()
        if key in stack:
            return source_name_map.get(key, "")

        stack2 = set(stack)
        stack2.add(key)
        base = str(source_name_map.get(key, "") or "").strip()
        formula = str(formula_by_key.get(key, "") or "").strip()
        if formula == "":
            memo[key] = base
            return memo[key]

        expr = _replace_formula_components_with_sources(
            formula=formula,
            resolve_component=lambda comp: resolve_reserving_source(comp, stack2),
            known_dataset_names=known_names,
            source_map=source_name_map,
        )
        memo[key] = expr if expr else base
        return memo[key]

    return resolve_reserving_source

def _quote_reserving_source_components(
    expr: Any,
    atomic_components: Optional[List[str]] = None,
) -> str:
    text = str(expr or "").strip()
    if not text:
        return ""

    protected_components: Dict[str, str] = {}
    protected_text = text
    unique_components = sorted(
        set([str(v or "").strip() for v in (atomic_components or []) if str(v or "").strip()]),
        key=len,
        reverse=True,
    )
    for component in unique_components:
        pattern = re.compile(
            rf"(?<![A-Za-z0-9_]){re.escape(component)}(?![A-Za-z0-9_])",
            flags=re.IGNORECASE,
        )

        def _protect_match(m: re.Match[str]) -> str:
            key = f"__RCOMP_{len(protected_components)}__"
            protected_components[key] = str(m.group(0) or "").strip()
            return key

        protected_text = pattern.sub(_protect_match, protected_text)

    tokens = re.findall(r'"[^"]*"|\(|\)|[+\-*/]|[^()+\-*/]+', protected_text)
    out_parts: List[str] = []
    for token in tokens:
        t = str(token or "")
        if not t:
            continue
        if t in ("(", ")"):
            out_parts.append(t)
            continue
        if t in ("+", "-", "*", "/"):
            out_parts.append(f" {t} ")
            continue

        component = t.strip()
        if not component:
            continue
        if component in protected_components:
            component = protected_components[component]
        if component.startswith('"') and component.endswith('"') and len(component) >= 2:
            component = component[1:-1].strip()
        component = component.replace('"', '\\"')
        out_parts.append(f'"{component}"')

    out = "".join(out_parts)
    out = re.sub(r"\s+", " ", out).strip()
    out = re.sub(r"\(\s+", "(", out)
    out = re.sub(r"\s+\)", ")", out)
    return out

def refresh_reserving_class_types_json(
    project_name: str,
    source_fields_override: Optional[List[Dict[str, Any]]] = None,
    rows_override: Optional[List[List[Any]]] = None,
) -> Dict[str, Any]:
    source_fields = source_fields_override if source_fields_override is not None else _load_reserving_class_value_fields(project_name)
    source_rows, source_names_lower, source_name_level_pairs = _build_reserving_class_source_rows(source_fields)
    previous_data = _load_reserving_class_types_raw_data(project_name)

    if rows_override is not None:
        rows_override_norm: List[List[str]] = []
        for raw in rows_override:
            if not isinstance(raw, list):
                continue
            name = str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip()
            level = str(raw[1] if len(raw) > 1 and raw[1] is not None else "").strip()
            formula = _normalize_formula_operator_spacing(raw[2] if len(raw) > 2 else "")
            eex_formula = _normalize_formula_operator_spacing(raw[3] if len(raw) > 3 else "")
            rows_override_norm.append([name, level, formula, eex_formula])
        changed_rows = _get_changed_reserving_class_formula_rows(
            previous_rows=previous_data.get("rows", []),
            next_rows=rows_override_norm,
            source_name_level_pairs=source_name_level_pairs,
        )
        _validate_reserving_class_formula_components(
            rows=changed_rows,
            source_rows=source_rows,
            known_rows=rows_override_norm,
        )
        base_data = normalize_reserving_class_types_data({
            "columns": list(RESERVING_CLASS_TYPES_COLUMNS),
            "rows": rows_override_norm,
        })
    else:
        base_data = previous_data

    merged_rows = _merge_reserving_class_types_rows(
        existing_rows=base_data.get("rows", []),
        source_rows=source_rows,
        source_name_level_pairs=source_name_level_pairs,
    )

    source_name_map: Dict[str, str] = {}
    for row in source_rows:
        if not isinstance(row, list):
            continue
        name = str(row[0] if len(row) > 0 and row[0] is not None else "").strip()
        if not name:
            continue
        source_name_map[_canon_reserving_class_type_name(name)] = name

    resolve_source = _build_reserving_class_type_source_resolver(
        reserving_rows=merged_rows,
        source_name_map=source_name_map,
    )
    source_component_candidates = []
    for row in merged_rows:
        if not isinstance(row, list):
            continue
        row_name = str(row[0] if len(row) > 0 and row[0] is not None else "").strip()
        if row_name:
            source_component_candidates.append(row_name)

    file_rows: List[List[str]] = []
    effective_source_names: List[str] = []
    seen_effective_source_keys: set[Tuple[str, str]] = set()
    for row in merged_rows:
        if not isinstance(row, list):
            continue
        name = str(row[0] if len(row) > 0 and row[0] is not None else "").strip()
        level = str(row[1] if len(row) > 1 and row[1] is not None else "").strip()
        formula = str(row[2] if len(row) > 2 and row[2] is not None else "").strip()
        eex_formula = str(row[3] if len(row) > 3 and row[3] is not None else "").strip()
        key = _canon_reserving_class_type_name(name)
        source_pair = (key, level)
        if source_pair in source_name_level_pairs and formula == "" and source_pair not in seen_effective_source_keys:
            seen_effective_source_keys.add(source_pair)
            effective_source_names.append(name)
        source = str(resolve_source(name) if name else "").strip()
        if not source and (source_pair in source_name_level_pairs or key in source_names_lower):
            source = name
        formula_atomic_components = _extract_formula_components(formula, source_component_candidates)
        atomic_components = list(source_component_candidates)
        if formula_atomic_components:
            atomic_components.extend(formula_atomic_components)
        source = _quote_reserving_source_components(source, atomic_components=atomic_components)
        file_rows.append([name, level, formula, eex_formula, source])

    payload = {
        "columns": list(RESERVING_CLASS_TYPES_FILE_COLUMNS),
        "rows": file_rows,
        "updated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
    }

    output_path = get_reserving_class_types_path(project_name)
    output_xlsx_path = _get_reserving_class_types_xlsx_path(output_path)
    should_write = True
    if os.path.exists(output_path):
        try:
            with open(output_path, "r", encoding="utf-8") as f:
                existing_raw = json.load(f)
            existing_norm = normalize_reserving_class_types_data(existing_raw)
            if (
                existing_norm.get("columns", []) == payload.get("columns", [])
                and existing_norm.get("rows", []) == payload.get("rows", [])
            ):
                should_write = False
                if isinstance(existing_raw, dict) and existing_raw.get("updated_at") is not None:
                    payload["updated_at"] = existing_raw.get("updated_at")
        except Exception:
            should_write = True
    if not should_write and not os.path.exists(output_xlsx_path):
        should_write = True

    write_out = {"json_path": output_path, "xlsx_path": output_xlsx_path}
    if should_write:
        write_out = save_reserving_class_types_payload(output_path, payload)

    return {
        "path": output_path,
        "xlsx_path": str(write_out.get("xlsx_path", "") or output_xlsx_path),
        "row_count": len(file_rows),
        "source_derived_count": len(effective_source_names),
        "source_derived_names": effective_source_names,
        "data": payload,
        "ui_data": _to_reserving_class_types_ui_data(payload),
    }


# ---------------------------------------------------------------------------
# Values and combinations
# ---------------------------------------------------------------------------

def _norm_path_for_compare(path_value: Any) -> str:
    p = str(path_value or "").strip()
    if not p:
        return ""
    try:
        return os.path.normcase(os.path.normpath(p))
    except Exception:
        return p.lower()

def _build_reserving_field_signature(field_defs: List[Dict[str, Any]]) -> str:
    canonical_items: List[str] = []
    for item in field_defs:
        name = str(item.get("field_name", "") or "").strip().lower()
        level = item.get("level")
        level_text = str(level) if isinstance(level, int) else ""
        if not name:
            continue
        canonical_items.append(f"{name}|{level_text}")
    canonical_items.sort()
    blob = "||".join(canonical_items)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()

def _payload_matches_reserving_snapshot(
    payload: Dict[str, Any],
    *,
    field_signature: str,
    table_path_norm: str,
    source_csv_mtime: Optional[float],
    source_csv_size: Optional[int],
) -> bool:
    old_sig = str(payload.get("field_signature", "") or "").strip()
    old_table = _norm_path_for_compare(payload.get("table_path", ""))
    if old_sig != field_signature or old_table != table_path_norm:
        return False

    if source_csv_mtime is None or source_csv_size is None:
        return True

    old_mtime: Optional[float] = None
    old_size: Optional[int] = None
    try:
        if payload.get("source_csv_mtime") is not None:
            old_mtime = float(payload.get("source_csv_mtime"))
    except Exception:
        old_mtime = None
    try:
        if payload.get("source_csv_size") is not None:
            old_size = int(payload.get("source_csv_size"))
    except Exception:
        old_size = None

    return (
        old_mtime is not None
        and abs(old_mtime - source_csv_mtime) <= 0.001
        and old_size is not None
        and old_size == source_csv_size
    )

def _load_table_summary_cached_meta(project_name: str, table_path: str) -> Tuple[Optional[float], List[str]]:
    tpath = str(table_path or "").strip()
    if not tpath:
        return (None, [])
    try:
        cache_path = get_cache_path(tpath, project_name=project_name)
    except Exception:
        return (None, [])
    if not os.path.exists(cache_path):
        return (None, [])
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return (None, [])
    if not isinstance(raw, dict):
        return (None, [])

    csv_mtime: Optional[float] = None
    try:
        m = raw.get("csv_mtime")
        if m is not None:
            csv_mtime = float(m)
    except Exception:
        csv_mtime = None

    names: List[str] = []
    seen: set[str] = set()
    columns = raw.get("columns", [])
    if isinstance(columns, list):
        for col in columns:
            if not isinstance(col, dict):
                continue
            name = str(col.get("name", "") or "").strip()
            if not name or name in seen:
                continue
            seen.add(name)
            names.append(name)
    return (csv_mtime, names)

def _extract_reserving_class_field_defs(rows: Any) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seen: set[str] = set()
    if not isinstance(rows, list):
        return out

    for row in rows:
        if not isinstance(row, dict):
            continue
        significance = str(row.get("significance", "") or "").strip()
        if significance != "Reserving Class":
            continue
        field_name = str(row.get("field_name", "") or "").strip()
        if not field_name:
            continue
        key = field_name.lower()
        if key in seen:
            continue
        seen.add(key)

        level: Optional[int] = None
        try:
            raw_level = row.get("level")
            if raw_level is not None and str(raw_level).strip() != "":
                n = int(raw_level)
                if n >= 1:
                    level = n
        except Exception:
            level = None

        out.append({"field_name": field_name, "level": level})

    out.sort(key=lambda x: ((x.get("level") if isinstance(x.get("level"), int) else 10**9), str(x.get("field_name", "")).lower()))
    return out

def _load_field_mapping_rows_and_table_path(project_name: str) -> Tuple[List[Dict[str, Any]], str]:
    try:
        filepath = get_field_mapping_path(project_name)
    except ValueError:
        return ([], "")
    if not os.path.exists(filepath):
        return ([], "")
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except Exception:
        return ([], "")
    rows = raw.get("rows", []) if isinstance(raw, dict) else []
    table_path = str(raw.get("table_path", "") or "").strip() if isinstance(raw, dict) else ""
    if not isinstance(rows, list):
        rows = []
    rows_norm: List[Dict[str, Any]] = []
    for row in rows:
        if isinstance(row, dict):
            rows_norm.append(row)
    return (rows_norm, table_path)

def _collect_distinct_values_from_csv(
    csv_path: str,
    field_names: List[str],
    available_columns: Optional[List[str]] = None,
    combination_field_names: Optional[List[str]] = None,
) -> Tuple[Dict[str, List[str]], List[str], float, int, List[str]]:
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")
    if not os.path.isfile(csv_path):
        raise ValueError(f"CSV path is not a file: {csv_path}")

    st = os.stat(csv_path)
    if isinstance(available_columns, list) and len(available_columns) > 0:
        available_cols = [str(c) for c in available_columns if str(c or "").strip()]
    else:
        header_df = pd.read_csv(csv_path, nrows=0)
        available_cols = [str(c) for c in list(header_df.columns)]
    available_set = set(available_cols)

    existing = [c for c in field_names if c in available_set]
    missing = [c for c in field_names if c not in available_set]
    values_by_field: Dict[str, List[str]] = {}
    combination_values: List[str] = []

    if existing:
        df = pd.read_csv(csv_path, usecols=existing, dtype=str, keep_default_na=False, low_memory=False)
        for col in existing:
            seen_vals: set[str] = set()
            vals: List[str] = []
            series = df[col]
            for raw in series.tolist():
                s = str(raw if raw is not None else "").strip()
                if not s:
                    continue
                if s in seen_vals:
                    continue
                seen_vals.add(s)
                vals.append(s)
            values_by_field[col] = vals

        combo_fields: List[str] = []
        if isinstance(combination_field_names, list) and len(combination_field_names) > 0:
            if all(str(c) in available_set for c in combination_field_names):
                combo_fields = [str(c) for c in combination_field_names]
        if combo_fields:
            seen_combo: set[str] = set()
            for row in df[combo_fields].itertuples(index=False, name=None):
                parts: List[str] = []
                skip = False
                for raw in row:
                    s = str(raw if raw is not None else "").strip()
                    if not s:
                        skip = True
                        break
                    parts.append(s)
                if skip:
                    continue
                combo = "\\".join(parts)
                if combo in seen_combo:
                    continue
                seen_combo.add(combo)
                combination_values.append(combo)

    return (values_by_field, missing, st.st_mtime, st.st_size, combination_values)

def _refresh_reserving_class_combinations_cache(
    project_name: str,
    *,
    table_path: str,
    field_defs: List[Dict[str, Any]],
    field_signature: str,
    source_csv_mtime: Optional[float],
    source_csv_size: Optional[int],
    missing_columns: Optional[List[str]] = None,
    combinations_override: Optional[List[str]] = None,
    available_columns: Optional[List[str]] = None,
    force: bool = False,
) -> Dict[str, Any]:
    output_path = get_reserving_class_combinations_path(project_name)
    table_path_norm = _norm_path_for_compare(table_path)
    field_names = [str(x.get("field_name", "") or "").strip() for x in field_defs if str(x.get("field_name", "") or "").strip()]

    existing_payload: Optional[Dict[str, Any]] = None
    if os.path.exists(output_path):
        try:
            with open(output_path, "r", encoding="utf-8") as f:
                raw_existing = json.load(f)
            if isinstance(raw_existing, dict):
                existing_payload = raw_existing
        except Exception:
            existing_payload = None

    existing_has_expanded_paths = (
        isinstance(existing_payload, dict)
        and (
            "aggregate_rule_signature" in existing_payload
            or "path_count" in existing_payload
            or "paths" in existing_payload
            or "tree" in existing_payload
            or "rule_count" in existing_payload
        )
    )

    if (not force) and existing_payload is not None and _payload_matches_reserving_snapshot(
        existing_payload,
        field_signature=field_signature,
        table_path_norm=table_path_norm,
        source_csv_mtime=source_csv_mtime,
        source_csv_size=source_csv_size,
    ) and (not existing_has_expanded_paths):
        cached_values = existing_payload.get("combinations", [])
        if not isinstance(cached_values, list):
            cached_values = []
        missing_cached = existing_payload.get("missing_columns", [])
        if not isinstance(missing_cached, list):
            missing_cached = []
        return {
            "combination_path": output_path,
            "combination_count": len(cached_values),
            "missing_columns": [str(x) for x in missing_cached if str(x or "").strip()],
            "combination_cached": True,
        }

    missing = [str(x) for x in (missing_columns or []) if str(x or "").strip()]
    combination_values: List[str] = []
    if field_names and not missing:
        if combinations_override is not None:
            seen: set[str] = set()
            for item in combinations_override:
                s = str(item if item is not None else "").strip()
                if not s or s in seen:
                    continue
                seen.add(s)
                combination_values.append(s)
        else:
            _, missing_from_csv, source_csv_mtime2, source_csv_size2, combos_from_csv = _collect_distinct_values_from_csv(
                table_path,
                field_names,
                available_columns=available_columns,
                combination_field_names=field_names,
            )
            if source_csv_mtime is None:
                source_csv_mtime = source_csv_mtime2
            if source_csv_size is None:
                source_csv_size = source_csv_size2
            if missing_from_csv:
                missing = [str(x) for x in missing_from_csv if str(x or "").strip()]
                combination_values = []
            else:
                combination_values = combos_from_csv

    fields_payload: List[Dict[str, Any]] = []
    for item in field_defs:
        fields_payload.append({
            "field_name": str(item.get("field_name", "") or "").strip(),
            "level": item.get("level"),
        })

    payload = {
        "project_name": project_name,
        "table_path": table_path,
        "updated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "source_csv_mtime": source_csv_mtime,
        "source_csv_size": source_csv_size,
        "field_signature": field_signature,
        "field_count": len(fields_payload),
        "missing_columns": missing,
        "fields": fields_payload,
        "combination_count": len(combination_values),
        "combinations": combination_values,
    }

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    tmp_path = output_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    os.replace(tmp_path, output_path)

    return {
        "combination_path": output_path,
        "combination_count": len(combination_values),
        "missing_columns": missing,
        "combination_cached": False,
    }

def refresh_reserving_class_values(
    project_name: str,
    table_path_override: Optional[str] = None,
    mapping_rows_override: Optional[List[Dict[str, Any]]] = None,
    force: bool = False,
) -> Dict[str, Any]:
    project_name = str(project_name or "").strip()
    if not project_name:
        raise ValueError("project_name is required")

    if mapping_rows_override is None:
        rows, table_path_from_mapping = _load_field_mapping_rows_and_table_path(project_name)
    else:
        rows = mapping_rows_override
        _, table_path_from_mapping = _load_field_mapping_rows_and_table_path(project_name)

    table_path = str(table_path_override or "").strip() or str(table_path_from_mapping or "").strip()
    field_defs = _extract_reserving_class_field_defs(rows)
    field_names = [str(x.get("field_name", "") or "").strip() for x in field_defs if str(x.get("field_name", "") or "").strip()]

    values_by_field: Dict[str, List[str]] = {}
    missing_columns: List[str] = []
    combinations: List[str] = []
    source_csv_mtime: Optional[float] = None
    source_csv_size: Optional[int] = None
    field_signature = _build_reserving_field_signature(field_defs)
    output_path = get_reserving_class_values_path(project_name)

    table_path_norm = _norm_path_for_compare(table_path)

    # Fast path: if field signature + source CSV metadata did not change, reuse existing JSON.
    existing_payload: Optional[Dict[str, Any]] = None
    if os.path.exists(output_path):
        try:
            with open(output_path, "r", encoding="utf-8") as f:
                raw_existing = json.load(f)
            if isinstance(raw_existing, dict):
                existing_payload = raw_existing
        except Exception:
            existing_payload = None

    summary_csv_mtime: Optional[float] = None
    summary_columns: List[str] = []
    if field_names:
        if not table_path:
            raise ValueError("Table path is required to refresh reserving class values.")

        st = os.stat(table_path)
        source_csv_mtime = st.st_mtime
        source_csv_size = st.st_size
        summary_csv_mtime, summary_columns = _load_table_summary_cached_meta(project_name, table_path)
        can_use_summary_cols = (
            summary_columns
            and summary_csv_mtime is not None
            and source_csv_mtime is not None
            and abs(float(summary_csv_mtime) - float(source_csv_mtime)) <= 0.001
        )

        if (not force) and existing_payload is not None and _payload_matches_reserving_snapshot(
            existing_payload,
            field_signature=field_signature,
            table_path_norm=table_path_norm,
            source_csv_mtime=source_csv_mtime,
            source_csv_size=source_csv_size,
        ):
            existing_fields = existing_payload.get("fields", [])
            if not isinstance(existing_fields, list):
                existing_fields = []
            value_count_cached = 0
            for item in existing_fields:
                if not isinstance(item, dict):
                    continue
                try:
                    value_count_cached += int(item.get("distinct_count", 0))
                except Exception:
                    vals = item.get("distinct_values", [])
                    if isinstance(vals, list):
                        value_count_cached += len(vals)
            missing_cached = existing_payload.get("missing_columns", [])
            if not isinstance(missing_cached, list):
                missing_cached = []
            combo_out = _refresh_reserving_class_combinations_cache(
                project_name,
                table_path=table_path,
                field_defs=field_defs,
                field_signature=field_signature,
                source_csv_mtime=source_csv_mtime,
                source_csv_size=source_csv_size,
                missing_columns=[str(x) for x in missing_cached if str(x or "").strip()],
                combinations_override=None,
                available_columns=(summary_columns if can_use_summary_cols else None),
                force=force,
            )
            rct_out = refresh_reserving_class_types_json(
                project_name,
                source_fields_override=existing_fields,
                rows_override=None,
            )
            return {
                "path": output_path,
                "field_count": len(existing_fields),
                "value_count": value_count_cached,
                "missing_columns": missing_cached,
                "table_path": table_path,
                "cached": True,
                "combination_path": combo_out.get("combination_path", ""),
                "combination_count": combo_out.get("combination_count", 0),
                "combination_cached": combo_out.get("combination_cached", False),
                "reserving_class_types_path": rct_out.get("path", ""),
                "reserving_class_types_count": rct_out.get("row_count", 0),
            }

        values_by_field, missing_columns, source_csv_mtime, source_csv_size, combinations = _collect_distinct_values_from_csv(
            table_path,
            field_names,
            available_columns=(summary_columns if can_use_summary_cols else None),
            combination_field_names=field_names,
        )
    else:
        # No reserving class fields: reuse existing payload if signature/table-path unchanged.
        if (not force) and existing_payload is not None and _payload_matches_reserving_snapshot(
            existing_payload,
            field_signature=field_signature,
            table_path_norm=table_path_norm,
            source_csv_mtime=None,
            source_csv_size=None,
        ):
            existing_fields = existing_payload.get("fields", [])
            if not isinstance(existing_fields, list):
                existing_fields = []
            combo_out = _refresh_reserving_class_combinations_cache(
                project_name,
                table_path=table_path,
                field_defs=field_defs,
                field_signature=field_signature,
                source_csv_mtime=None,
                source_csv_size=None,
                missing_columns=[],
                combinations_override=[],
                available_columns=None,
                force=force,
            )
            rct_out = refresh_reserving_class_types_json(
                project_name,
                source_fields_override=existing_fields,
                rows_override=None,
            )
            return {
                "path": output_path,
                "field_count": len(existing_fields),
                "value_count": 0,
                "missing_columns": [],
                "table_path": table_path,
                "cached": True,
                "combination_path": combo_out.get("combination_path", ""),
                "combination_count": combo_out.get("combination_count", 0),
                "combination_cached": combo_out.get("combination_cached", False),
                "reserving_class_types_path": rct_out.get("path", ""),
                "reserving_class_types_count": rct_out.get("row_count", 0),
            }

    fields_payload: List[Dict[str, Any]] = []
    total_values = 0
    for item in field_defs:
        field_name = str(item.get("field_name", "") or "").strip()
        values = list(values_by_field.get(field_name, []))
        total_values += len(values)
        fields_payload.append({
            "field_name": field_name,
            "level": item.get("level"),
            "distinct_count": len(values),
            "distinct_values": values,
        })

    payload = {
        "project_name": project_name,
        "table_path": table_path,
        "updated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "source_csv_mtime": source_csv_mtime,
        "source_csv_size": source_csv_size,
        "field_signature": field_signature,
        "field_count": len(fields_payload),
        "missing_columns": missing_columns,
        "fields": fields_payload,
    }

    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    tmp_path = output_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    os.replace(tmp_path, output_path)

    combo_out = _refresh_reserving_class_combinations_cache(
        project_name,
        table_path=table_path,
        field_defs=field_defs,
        field_signature=field_signature,
        source_csv_mtime=source_csv_mtime,
        source_csv_size=source_csv_size,
        missing_columns=missing_columns,
        combinations_override=combinations,
        available_columns=(summary_columns if summary_columns else None),
        force=force,
    )
    rct_out = refresh_reserving_class_types_json(
        project_name,
        source_fields_override=fields_payload,
        rows_override=None,
    )

    return {
        "path": output_path,
        "field_count": len(fields_payload),
        "value_count": total_values,
        "missing_columns": missing_columns,
        "table_path": table_path,
        "cached": False,
        "combination_path": combo_out.get("combination_path", ""),
        "combination_count": combo_out.get("combination_count", 0),
        "combination_cached": combo_out.get("combination_cached", False),
        "reserving_class_types_path": rct_out.get("path", ""),
        "reserving_class_types_count": rct_out.get("row_count", 0),
    }


# ---------------------------------------------------------------------------
# Path tree hierarchy
# ---------------------------------------------------------------------------

def _extract_reserving_class_level_rules(
    reserving_rows: List[List[Any]],
) -> Tuple[Dict[int, Dict[str, List[str]]], int, Dict[int, Dict[str, List[str]]]]:
    rows_by_level: Dict[int, List[Dict[str, str]]] = {}
    for raw in reserving_rows:
        if not isinstance(raw, list):
            continue
        name = str(raw[0] if len(raw) > 0 and raw[0] is not None else "").strip()
        if not name:
            continue
        level = _parse_positive_int(raw[1] if len(raw) > 1 else None)
        if level is None:
            continue
        formula = _normalize_formula_operator_spacing(raw[2] if len(raw) > 2 else "")
        rows_by_level.setdefault(level, []).append({
            "name": name,
            "formula": formula,
        })

    component_to_parent_names_by_level: Dict[int, Dict[str, List[str]]] = {}
    parent_to_components_by_level: Dict[int, Dict[str, List[str]]] = {}
    rule_count = 0
    for level, rows in rows_by_level.items():
        known_names = [str(r.get("name", "") or "").strip() for r in rows]
        level_map: Dict[str, List[str]] = {}
        level_parent_components: Dict[str, List[str]] = {}
        for row in rows:
            name = str(row.get("name", "") or "").strip()
            formula = str(row.get("formula", "") or "").strip()
            if not name or not formula:
                continue
            parent_key = _canon_reserving_class_type_name(name)
            if not parent_key:
                continue
            components = _extract_formula_components(formula, known_names)
            if not components:
                continue
            if parent_key not in level_parent_components:
                level_parent_components[parent_key] = []
            used_components: set[str] = set()
            for comp in components:
                comp_key = _canon_reserving_class_type_name(comp)
                if not comp_key or comp_key == parent_key or comp_key in used_components:
                    continue
                used_components.add(comp_key)
                if all(_canon_reserving_class_type_name(v) != comp_key for v in level_parent_components[parent_key]):
                    level_parent_components[parent_key].append(comp)
                if comp_key not in level_map:
                    level_map[comp_key] = []
                if all(_canon_reserving_class_type_name(v) != parent_key for v in level_map[comp_key]):
                    level_map[comp_key].append(name)
            if used_components:
                rule_count += 1
        if level_map:
            component_to_parent_names_by_level[level] = level_map
        if level_parent_components:
            parent_to_components_by_level[level] = level_parent_components

    return component_to_parent_names_by_level, rule_count, parent_to_components_by_level

def _expand_reserving_class_name_for_level(
    name: str,
    component_to_parent_names: Dict[str, List[str]],
) -> List[str]:
    start = str(name or "").strip()
    if not start:
        return []

    out: List[str] = []
    seen: set[str] = set()
    queue: List[str] = [start]
    idx = 0
    while idx < len(queue):
        current = str(queue[idx] or "").strip()
        idx += 1
        current_key = _canon_reserving_class_type_name(current)
        if not current_key or current_key in seen:
            continue
        seen.add(current_key)
        out.append(current)
        for parent_name in component_to_parent_names.get(current_key, []):
            p = str(parent_name or "").strip()
            p_key = _canon_reserving_class_type_name(p)
            if not p or not p_key or p_key in seen:
                continue
            queue.append(p)
    return out

def _split_reserving_combo_parts(raw: Any) -> List[str]:
    parts = [str(p).strip() for p in str(raw if raw is not None else "").split("\\")]
    if not parts or any(not p for p in parts):
        return []
    return parts

def _build_reserving_descendants_lookup_by_level(
    parent_to_components_by_level: Dict[int, Dict[str, List[str]]],
) -> Dict[int, Dict[str, set[str]]]:
    out: Dict[int, Dict[str, set[str]]] = {}
    for level, parent_map_raw in parent_to_components_by_level.items():
        parent_map: Dict[str, List[str]] = {}
        for parent_name, comps in parent_map_raw.items():
            p_key = _canon_reserving_class_type_name(parent_name)
            if not p_key:
                continue
            if p_key not in parent_map:
                parent_map[p_key] = []
            for comp in comps or []:
                c = str(comp if comp is not None else "").strip()
                c_key = _canon_reserving_class_type_name(c)
                if not c_key:
                    continue
                if all(_canon_reserving_class_type_name(v) != c_key for v in parent_map[p_key]):
                    parent_map[p_key].append(c)

        memo: Dict[str, set[str]] = {}

        def _descendants(parent_key: str, stack: Optional[set[str]] = None) -> set[str]:
            if parent_key in memo:
                return set(memo[parent_key])
            if stack is None:
                stack = set()
            if parent_key in stack:
                return {parent_key}

            stack2 = set(stack)
            stack2.add(parent_key)
            acc: set[str] = {parent_key}
            for comp_name in parent_map.get(parent_key, []):
                comp_key = _canon_reserving_class_type_name(comp_name)
                if not comp_key:
                    continue
                acc.add(comp_key)
                if comp_key in parent_map:
                    acc.update(_descendants(comp_key, stack2))
            memo[parent_key] = set(acc)
            return set(acc)

        lookup: Dict[str, set[str]] = {}
        for parent_key in parent_map.keys():
            lookup[parent_key] = _descendants(parent_key)
        out[level] = lookup
    return out

def _build_path_tree_structure(paths: List[str], level_labels: List[str]) -> Dict[str, Any]:
    root: Dict[str, Any] = {
        "name": "All",
        "path": "",
        "level_index": 0,
        "level_label": "All",
        "_children": {},
    }

    for raw in paths:
        path = str(raw if raw is not None else "").strip()
        if not path:
            continue
        parts = [str(p).strip() for p in path.split("\\") if str(p).strip()]
        if not parts:
            continue
        cur = root
        acc: List[str] = []
        for idx, part in enumerate(parts):
            acc.append(part)
            children = cur["_children"]
            if part not in children:
                children[part] = {
                    "name": part,
                    "path": "\\".join(acc),
                    "level_index": idx + 1,
                    "level_label": level_labels[idx] if idx < len(level_labels) else f"Level {idx + 1}",
                    "_children": {},
                }
            cur = children[part]

    def _finalize(node: Dict[str, Any]) -> Dict[str, Any]:
        kids_map = node.get("_children", {})
        kid_keys = sorted(kids_map.keys(), key=lambda k: str(k).lower())
        kids = [_finalize(kids_map[k]) for k in kid_keys]
        return {
            "name": node.get("name", ""),
            "path": node.get("path", ""),
            "level_index": node.get("level_index", 0),
            "level_label": node.get("level_label", ""),
            "children": kids,
        }

    return _finalize(root)

def _extract_reserving_base_paths_from_combo_payload(combo_payload: Dict[str, Any]) -> List[str]:
    base_paths_raw = combo_payload.get("combinations", []) if isinstance(combo_payload, dict) else []
    base_paths: List[str] = []
    seen_base: set[str] = set()
    for raw in base_paths_raw if isinstance(base_paths_raw, list) else []:
        s = str(raw if raw is not None else "").strip()
        if not s or s in seen_base:
            continue
        seen_base.add(s)
        base_paths.append(s)
    return base_paths

def _extract_reserving_level_info_from_combo_payload(
    combo_payload: Dict[str, Any],
    base_paths: List[str],
) -> Tuple[List[int], List[str], List[Dict[str, Any]]]:
    fields_raw = combo_payload.get("fields", []) if isinstance(combo_payload, dict) else []
    level_numbers: List[int] = []
    level_labels: List[str] = []
    levels_payload: List[Dict[str, Any]] = []
    if isinstance(fields_raw, list):
        for idx, item in enumerate(fields_raw):
            field_name = str((item or {}).get("field_name", "") if isinstance(item, dict) else "").strip()
            level_num = _parse_positive_int((item or {}).get("level") if isinstance(item, dict) else None)
            if level_num is None:
                level_num = idx + 1
            level_numbers.append(level_num)
            label = field_name or f"Level {idx + 1}"
            level_labels.append(label)
            levels_payload.append({
                "index": idx + 1,
                "level": level_num,
                "field_name": field_name,
            })

    if not level_numbers and base_paths:
        sample_parts = [str(p).strip() for p in base_paths[0].split("\\") if str(p).strip()]
        for idx in range(len(sample_parts)):
            level_numbers.append(idx + 1)
            level_labels.append(f"Level {idx + 1}")
            levels_payload.append({
                "index": idx + 1,
                "level": idx + 1,
                "field_name": "",
            })

    return (level_numbers, level_labels, levels_payload)

def _canonical_reserving_parts_key(parts: List[str]) -> str:
    if not parts:
        return ""
    keys: List[str] = []
    for raw in parts:
        key = _canon_reserving_class_type_name(raw)
        if not key:
            return ""
        keys.append(key)
    return "\\".join(keys)

def _canonicalize_reserving_name_list(names: List[str]) -> List[str]:
    sorted_names = sorted(
        [str(v if v is not None else "").strip() for v in names if str(v if v is not None else "").strip()],
        key=lambda x: x.lower(),
    )
    out: List[str] = []
    seen: set[str] = set()
    for name in sorted_names:
        key = _canon_reserving_class_type_name(name)
        if not key or key in seen:
            continue
        seen.add(key)
        out.append(name)
    return out

def _normalize_reserving_children_cache(raw: Any) -> Dict[str, List[str]]:
    out: Dict[str, List[str]] = {}
    if not isinstance(raw, dict):
        return out
    for raw_key, raw_names in raw.items():
        key = str(raw_key if raw_key is not None else "")
        if not isinstance(raw_names, list):
            continue
        names = _canonicalize_reserving_name_list([str(v if v is not None else "") for v in raw_names])
        out[key] = names
    return out

def _is_same_optional_mtime(a: Any, b: Any) -> bool:
    if a is None and b is None:
        return True
    if a is None or b is None:
        return False
    try:
        return abs(float(a) - float(b)) <= 0.001
    except Exception:
        return False

def _build_reserving_level_display_lookup_by_level(
    base_parts_list: List[List[str]],
    level_numbers: List[int],
    rules_by_level: Dict[int, Dict[str, List[str]]],
    parent_to_components_by_level: Dict[int, Dict[str, List[str]]],
) -> Dict[int, Dict[str, str]]:
    out: Dict[int, Dict[str, str]] = {}

    for parts in base_parts_list:
        for idx, raw in enumerate(parts):
            level_num = level_numbers[idx] if idx < len(level_numbers) else (idx + 1)
            key = _canon_reserving_class_type_name(raw)
            value = str(raw if raw is not None else "").strip()
            if not key or not value:
                continue
            if level_num not in out:
                out[level_num] = {}
            if key not in out[level_num]:
                out[level_num][key] = value

    for level_num, parent_map in parent_to_components_by_level.items():
        if level_num not in out:
            out[level_num] = {}
        for parent_key, components in parent_map.items():
            p_key = _canon_reserving_class_type_name(parent_key)
            if p_key and p_key not in out[level_num]:
                out[level_num][p_key] = str(parent_key if parent_key is not None else "").strip() or p_key
            for comp in components or []:
                comp_name = str(comp if comp is not None else "").strip()
                c_key = _canon_reserving_class_type_name(comp_name)
                if not c_key or not comp_name:
                    continue
                if c_key not in out[level_num]:
                    out[level_num][c_key] = comp_name

    for level_num, comp_to_parents in rules_by_level.items():
        if level_num not in out:
            out[level_num] = {}
        for comp_key, parent_names in comp_to_parents.items():
            c_key = _canon_reserving_class_type_name(comp_key)
            if c_key and c_key not in out[level_num]:
                out[level_num][c_key] = str(comp_key if comp_key is not None else "").strip() or c_key
            for parent_name in parent_names:
                p_name = str(parent_name if parent_name is not None else "").strip()
                p_key = _canon_reserving_class_type_name(p_name)
                if not p_key or not p_name:
                    continue
                # Prefer the display casing from the explicit rule name.
                out[level_num][p_key] = p_name

    return out

def _expand_reserving_keys_with_parents_for_level(
    seed_keys: set[str],
    level_rules: Dict[str, List[str]],
) -> set[str]:
    out: set[str] = set(seed_keys)
    queue: List[str] = list(seed_keys)
    idx = 0
    while idx < len(queue):
        current_key = queue[idx]
        idx += 1
        for parent_name in level_rules.get(current_key, []):
            parent_key = _canon_reserving_class_type_name(parent_name)
            if not parent_key or parent_key in out:
                continue
            out.add(parent_key)
            queue.append(parent_key)
    return out

def _load_reserving_path_tree_cache_payload(cache_path: str) -> Dict[str, Any]:
    if not os.path.exists(cache_path):
        return {}
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        if isinstance(raw, dict):
            return raw
    except Exception:
        return {}
    return {}

def _write_reserving_path_tree_cache_payload(cache_path: str, payload: Dict[str, Any]) -> None:
    os.makedirs(os.path.dirname(cache_path), exist_ok=True)
    tmp_path = cache_path + ".tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)
    os.replace(tmp_path, cache_path)

def _build_reserving_child_nodes(
    child_names: List[str],
    prefix_parts_display: List[str],
    child_level_index: int,
    child_level_label: str,
    has_children: bool,
) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for child_name in child_names:
        parts = prefix_parts_display + [child_name]
        out.append({
            "name": child_name,
            "path": "\\".join(parts),
            "level_index": child_level_index,
            "level_label": child_level_label,
            "has_children": has_children,
        })
    return out

def get_reserving_class_path_tree_children(
    project_name: str,
    prefix: str = "",
    table_path_override: str = "",
    force: bool = False,
) -> Dict[str, Any]:
    project_name_clean = str(project_name or "").strip()
    if not project_name_clean:
        raise ValueError("project_name is required")

    refresh_out = refresh_reserving_class_values(
        project_name=project_name_clean,
        table_path_override=str(table_path_override or "").strip(),
        mapping_rows_override=None,
        force=bool(force),
    )

    combinations_path = str(refresh_out.get("combination_path") or "").strip()
    if not combinations_path:
        combinations_path = get_reserving_class_combinations_path(project_name_clean)

    combo_payload: Dict[str, Any] = {}
    if os.path.exists(combinations_path):
        with open(combinations_path, "r", encoding="utf-8") as f:
            raw_combo = json.load(f)
        if isinstance(raw_combo, dict):
            combo_payload = raw_combo

    base_paths = _extract_reserving_base_paths_from_combo_payload(combo_payload)
    level_numbers, level_labels, levels_payload = _extract_reserving_level_info_from_combo_payload(
        combo_payload,
        base_paths,
    )

    prefix_norm = _norm_tree_path(prefix)
    prefix_parts: List[str] = []
    if prefix_norm:
        prefix_parts = _split_reserving_combo_parts(prefix_norm)
        if not prefix_parts:
            raise ValueError("Invalid reserving class prefix path.")

    prefix_keys: List[str] = []
    for part in prefix_parts:
        key = _canon_reserving_class_type_name(part)
        if not key:
            raise ValueError("Invalid reserving class prefix path.")
        prefix_keys.append(key)

    child_level_index = len(prefix_parts) + 1
    child_level_label = level_labels[child_level_index - 1] if child_level_index - 1 < len(level_labels) else f"Level {child_level_index}"
    has_children = child_level_index < len(level_numbers)

    reserving_raw = _load_reserving_class_types_raw_data(project_name_clean)
    reserving_rows = reserving_raw.get("rows", []) if isinstance(reserving_raw, dict) else []
    rules_by_level, rule_count, parent_to_components_by_level = _extract_reserving_class_level_rules(
        reserving_rows if isinstance(reserving_rows, list) else [],
    )

    rct_path = get_reserving_class_types_path(project_name_clean)
    rct_mtime: Optional[float] = None
    try:
        if os.path.exists(rct_path):
            rct_mtime = os.stat(rct_path).st_mtime
    except Exception:
        rct_mtime = None

    cache_path = get_reserving_class_path_tree_path(project_name_clean)
    field_signature = str(combo_payload.get("field_signature", "") if isinstance(combo_payload, dict) else "").strip()
    table_path = str(combo_payload.get("table_path", "") if isinstance(combo_payload, dict) else "").strip()
    table_path_norm = _norm_path_for_compare(table_path)
    source_csv_mtime = combo_payload.get("source_csv_mtime") if isinstance(combo_payload, dict) else None
    source_csv_size = combo_payload.get("source_csv_size") if isinstance(combo_payload, dict) else None
    prefix_key = _canonical_reserving_parts_key(prefix_parts) if prefix_parts else ""

    existing_payload: Dict[str, Any] = {}
    cache_valid = False
    children_cache: Dict[str, List[str]] = {}
    existing_paths_raw: List[str] = []
    with _RESERVING_CLASS_PATH_TREE_LOCK:
        existing_payload = _load_reserving_path_tree_cache_payload(cache_path)
        if existing_payload and _payload_matches_reserving_snapshot(
            existing_payload,
            field_signature=field_signature,
            table_path_norm=table_path_norm,
            source_csv_mtime=source_csv_mtime,
            source_csv_size=source_csv_size,
        ) and _is_same_optional_mtime(existing_payload.get("reserving_class_types_mtime"), rct_mtime):
            cache_valid = True
            children_cache = _normalize_reserving_children_cache(existing_payload.get("children_cache"))
            existing_paths = existing_payload.get("paths", [])
            if isinstance(existing_paths, list):
                existing_paths_raw = [str(v if v is not None else "").strip() for v in existing_paths]

    if (
        not force
        and cache_valid
        and prefix_key in children_cache
    ):
        child_names_cached = children_cache.get(prefix_key, [])
        child_nodes_cached = _build_reserving_child_nodes(
            child_names=child_names_cached,
            prefix_parts_display=prefix_parts,
            child_level_index=child_level_index,
            child_level_label=child_level_label,
            has_children=has_children,
        )
        return {
            "project_name": project_name_clean,
            "prefix": prefix_norm,
            "levels": levels_payload,
            "child_level_index": child_level_index,
            "child_level_label": child_level_label,
            "children": child_nodes_cached,
            "from_cache": True,
            "cache_path": cache_path,
            "combination_path": combinations_path,
            "table_path": table_path,
            "base_path_count": len(base_paths),
            "generated_path_count": len(existing_paths_raw),
            "rule_count": rule_count,
            "cached_prefix_count": len(children_cache),
        }

    child_names: List[str] = []
    prefix_parts_display = list(prefix_parts)
    if child_level_index <= len(level_numbers):
        base_parts_list: List[List[str]] = []
        base_key_rows: List[List[str]] = []
        for raw in base_paths:
            parts = _split_reserving_combo_parts(raw)
            if not parts:
                continue
            keys: List[str] = []
            skip_row = False
            for part in parts:
                key = _canon_reserving_class_type_name(part)
                if not key:
                    skip_row = True
                    break
                keys.append(key)
            if skip_row:
                continue
            base_parts_list.append(parts)
            base_key_rows.append(keys)

        display_by_level = _build_reserving_level_display_lookup_by_level(
            base_parts_list,
            level_numbers,
            rules_by_level,
            parent_to_components_by_level,
        )

        for idx, part in enumerate(prefix_parts):
            level_num = level_numbers[idx] if idx < len(level_numbers) else (idx + 1)
            key = prefix_keys[idx] if idx < len(prefix_keys) else ""
            if not key:
                continue
            if level_num not in display_by_level:
                display_by_level[level_num] = {}
            if key not in display_by_level[level_num]:
                display_by_level[level_num][key] = part
            prefix_parts_display[idx] = display_by_level[level_num].get(key, part)

        descendants_lookup = _build_reserving_descendants_lookup_by_level(parent_to_components_by_level)
        matched_child_raw_keys: set[str] = set()
        child_idx0 = child_level_index - 1
        for row_keys in base_key_rows:
            if len(row_keys) <= child_idx0:
                continue
            matched = True
            for idx, prefix_key_part in enumerate(prefix_keys):
                if idx >= len(row_keys):
                    matched = False
                    break
                level_num = level_numbers[idx] if idx < len(level_numbers) else (idx + 1)
                parent_map = parent_to_components_by_level.get(level_num, {})
                if prefix_key_part in parent_map:
                    allowed_values = descendants_lookup.get(level_num, {}).get(prefix_key_part, {prefix_key_part})
                    if row_keys[idx] not in allowed_values:
                        matched = False
                        break
                else:
                    if row_keys[idx] != prefix_key_part:
                        matched = False
                        break
            if not matched:
                continue
            matched_child_raw_keys.add(row_keys[child_idx0])

        child_level_num = level_numbers[child_idx0]
        expanded_child_keys = _expand_reserving_keys_with_parents_for_level(
            matched_child_raw_keys,
            rules_by_level.get(child_level_num, {}),
        )
        child_display = display_by_level.get(child_level_num, {})
        child_names = _canonicalize_reserving_name_list([child_display.get(k, k) for k in expanded_child_keys])

    child_nodes = _build_reserving_child_nodes(
        child_names=child_names,
        prefix_parts_display=prefix_parts_display,
        child_level_index=child_level_index,
        child_level_label=child_level_label,
        has_children=has_children,
    )

    paths_for_cache: List[str] = []
    seen_path_keys: set[str] = set()
    for raw in existing_paths_raw:
        parts = _split_reserving_combo_parts(raw)
        key = _canonical_reserving_parts_key(parts)
        if not key or key in seen_path_keys:
            continue
        seen_path_keys.add(key)
        paths_for_cache.append("\\".join(parts))

    new_paths_added = 0
    for node in child_nodes:
        path = str(node.get("path", "") or "").strip()
        parts = _split_reserving_combo_parts(path)
        key = _canonical_reserving_parts_key(parts)
        if not key or key in seen_path_keys:
            continue
        if len(paths_for_cache) >= RESERVING_CLASS_PATH_TREE_MAX_GENERATED:
            break
        seen_path_keys.add(key)
        paths_for_cache.append(path)
        new_paths_added += 1

    existing_cached_child_names = children_cache.get(prefix_key) if cache_valid else None
    children_cache[prefix_key] = child_names

    root_tree = {"name": "All", "path": "", "level_index": 0, "level_label": "All", "children": []}
    existing_tree = existing_payload.get("tree")
    if isinstance(existing_tree, dict):
        root_tree = existing_tree
    if len(paths_for_cache) <= 20000:
        root_tree = _build_path_tree_structure(paths_for_cache, level_labels)

    payload_out = {
        "project_name": project_name_clean,
        "mode": "lazy",
        "table_path": table_path,
        "updated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "source_csv_mtime": source_csv_mtime,
        "source_csv_size": source_csv_size,
        "field_signature": field_signature,
        "levels": levels_payload,
        "rule_count": rule_count,
        "base_path_count": len(base_paths),
        "generated_path_count": len(paths_for_cache),
        "paths": paths_for_cache,
        "children_cache": children_cache,
        "tree": root_tree,
        "reserving_class_types_path": rct_path,
        "reserving_class_types_mtime": rct_mtime,
        "combination_path": combinations_path,
    }

    should_write = (
        force
        or (not cache_valid)
        or (new_paths_added > 0)
        or (existing_cached_child_names != child_names)
    )

    if should_write:
        with _RESERVING_CLASS_PATH_TREE_LOCK:
            latest_payload = _load_reserving_path_tree_cache_payload(cache_path)
            latest_valid = latest_payload and _payload_matches_reserving_snapshot(
                latest_payload,
                field_signature=field_signature,
                table_path_norm=table_path_norm,
                source_csv_mtime=source_csv_mtime,
                source_csv_size=source_csv_size,
            ) and _is_same_optional_mtime(latest_payload.get("reserving_class_types_mtime"), rct_mtime)

            latest_children_cache = _normalize_reserving_children_cache(latest_payload.get("children_cache")) if latest_valid else {}
            latest_children_cache[prefix_key] = child_names

            latest_paths_raw = latest_payload.get("paths", []) if latest_valid else []
            latest_paths: List[str] = []
            latest_seen: set[str] = set()
            for raw in latest_paths_raw if isinstance(latest_paths_raw, list) else []:
                parts = _split_reserving_combo_parts(raw)
                key = _canonical_reserving_parts_key(parts)
                if not key or key in latest_seen:
                    continue
                latest_seen.add(key)
                latest_paths.append("\\".join(parts))
            for node in child_nodes:
                path = str(node.get("path", "") or "").strip()
                parts = _split_reserving_combo_parts(path)
                key = _canonical_reserving_parts_key(parts)
                if not key or key in latest_seen:
                    continue
                if len(latest_paths) >= RESERVING_CLASS_PATH_TREE_MAX_GENERATED:
                    break
                latest_seen.add(key)
                latest_paths.append(path)

            latest_tree = latest_payload.get("tree") if isinstance(latest_payload.get("tree"), dict) else root_tree
            if len(latest_paths) <= 20000:
                latest_tree = _build_path_tree_structure(latest_paths, level_labels)

            payload_out = {
                "project_name": project_name_clean,
                "mode": "lazy",
                "table_path": table_path,
                "updated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
                "source_csv_mtime": source_csv_mtime,
                "source_csv_size": source_csv_size,
                "field_signature": field_signature,
                "levels": levels_payload,
                "rule_count": rule_count,
                "base_path_count": len(base_paths),
                "generated_path_count": len(latest_paths),
                "paths": latest_paths,
                "children_cache": latest_children_cache,
                "tree": latest_tree,
                "reserving_class_types_path": rct_path,
                "reserving_class_types_mtime": rct_mtime,
                "combination_path": combinations_path,
            }
            _write_reserving_path_tree_cache_payload(cache_path, payload_out)
            paths_for_cache = latest_paths
            children_cache = latest_children_cache

    return {
        "project_name": project_name_clean,
        "prefix": prefix_norm,
        "levels": levels_payload,
        "child_level_index": child_level_index,
        "child_level_label": child_level_label,
        "children": child_nodes,
        "from_cache": False,
        "cache_path": cache_path,
        "combination_path": combinations_path,
        "table_path": table_path,
        "base_path_count": len(base_paths),
        "generated_path_count": len(paths_for_cache),
        "rule_count": rule_count,
        "cached_prefix_count": len(children_cache),
    }
